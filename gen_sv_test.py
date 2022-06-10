import re
import math
import networkx as nx
from openpyxl import load_workbook
import math
import os
from colorama import Fore, Style
import shutil
import multiprocessing as mp
from itertools import permutations, combinations
import numpy as np
import time
import pickle
import json
from difflib import SequenceMatcher

########################## global settings ##########################
SIGNATURE = '/*********** Data Phase Auto Gen Tools ***********/\n'
COMPONENT_NAME_COLUMN_NAME = '主圖形名稱'
INPUT_DIR   = 'input'
ENV_DIR     = 'data_phase_env'
DATA_DIR    = 'data'
PATTERN_DIR = 'audio_data_phase_auto_gen_patterns'
CHECK_LOG_DIR = 'check_log'
NULL_NODE_ID = -1
TAB = '    '
SELECT_SIGNAL_RESERVED = 'reserved'
REG_READ_WRITE_CMD_ADDR = 'UVM_ADDR'
REG_READ_WRITE_CMD_DATA = 'UVM_DATA'
REG_READ_WRITE_VARIABLE = 'UVM_VARIABLE'
REG_READ_WRITE_VARIABLE_NAME = 'data_tmp'
RESERVED_SELECT_SIGNAL = SELECT_SIGNAL_RESERVED
OFFSET_SELECT_SIGNAL   = '_offset_'
FOR_SD_CHECK_ONLY = False
cust_audio_data_phase_coverage_model_changed = False

########################## global settings ##########################

def contents_append_endl(CONTENTS, tmp):
    CONTENTS.append(tmp + '\n')

def contents_append_tab_endl(CONTENTS, tmp):
    CONTENTS.append(TAB + tmp + '\n')

def contents_append_tab_tab_endl(CONTENTS, tmp):
    CONTENTS.append(TAB + TAB + tmp + '\n')

def ERROR_LOG(LOG):
    return '[ERROR] ' + LOG

def COLOR_MESSAGE(message, SETTINGS):
    #init()
    #print(SETTINGS + message + Style.RESET_ALL)
    print(message)

def ERROR_MESSAGE(message):
    SETTINGS = Style.BRIGHT + Fore.RED
    COLOR_MESSAGE(ERROR_LOG(message), SETTINGS)

def get_nodes_by_type(components_info, type):
    nodes = []
    for component_info in components_info:
        if component_info['Type'] == type:
            nodes.append(component_info)
            
    return nodes

def file_with_dir_path(output_file_name, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, output_file_name)

def write_contents_to_file(CONTENTS, output_file_name, output_dir, IS_APPEND_MODE = False, ENDL = '\n'):
    output_file_name = file_with_dir_path(output_file_name, output_dir)
   
    if IS_APPEND_MODE:
        output_file = open(output_file_name, "a", encoding = "UTF-8")
    else:
        output_file = open(output_file_name, "w", encoding = "UTF-8")

    for content in CONTENTS:
        output_file.write(content + ENDL)
    output_file.close()

def gen_coverage_data_phase(pn2_has_edge_covered, output_file_name = 'cust_audio_data_phase_coverage_model.sv', output_dir = ENV_DIR):
    CONTENTS  = []
    contents_append_endl(CONTENTS, SIGNATURE) 
    contents_append_endl(CONTENTS, '`ifndef CUST_AUDIO_DATA_PHASE_COVERAGE_MODEL__SV')
    contents_append_endl(CONTENTS, '`define CUST_AUDIO_DATA_PHASE_COVERAGE_MODEL__SV')
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, 'class cust_audio_data_phase_coverage_model extends audio_data_phase_coverage_model #(cust_audio_data_path_dump_mux_transaction, cust_audio_data_path_transition_model);')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, '`uvm_component_utils(cust_audio_data_phase_coverage_model)')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, 'covergroup data_phase_combination;')
    contents_append_tab_tab_endl(CONTENTS, 'option.per_instance = 1;')
    contents_append_tab_tab_endl(CONTENTS, 'coverpoint combination_idx')
    contents_append_tab_tab_endl(CONTENTS, '{')
    
    for i in range(len(pn2_has_edge_covered)):
        temp_str = '  bins C_{}_{} = '.format(pn2_has_edge_covered[i][0], pn2_has_edge_covered[i][1])
        temp_str += '{' + str(i) + '};'
        contents_append_tab_tab_endl(CONTENTS, temp_str)

    contents_append_tab_tab_endl(CONTENTS, '}')
    contents_append_tab_endl(CONTENTS, 'endgroup')
    contents_append_endl(CONTENTS, '')
        
    contents_append_tab_endl(CONTENTS, 'extern function new(string name = "cust_audio_data_phase_coverage_model", uvm_component parent);')
    contents_append_tab_endl(CONTENTS, 'extern virtual function void sample_cov();')
    contents_append_tab_endl(CONTENTS, 'extern virtual function real get_cov();')
    contents_append_endl(CONTENTS, 'endclass: cust_audio_data_phase_coverage_model')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '/** new function for cust_audio_data_phase_coverage_model */')
    contents_append_endl(CONTENTS, 'function cust_audio_data_phase_coverage_model::new(string name = "cust_audio_data_phase_coverage_model", uvm_component parent);')
    contents_append_tab_endl(CONTENTS, 'super.new(name, parent);')
    contents_append_tab_endl(CONTENTS, 'data_phase_combination = new();')
    contents_append_endl(CONTENTS, 'endfunction: new')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '/** sample cov for cust_audio_data_phase_coverage_model */')
    contents_append_endl(CONTENTS, 'function void cust_audio_data_phase_coverage_model::sample_cov();')
    contents_append_tab_endl(CONTENTS, 'for(int i=0; i<$size(cfg.combination_idx); i++) begin')
    contents_append_tab_endl(CONTENTS, '  combination_idx = cfg.combination_idx[i];')
    contents_append_tab_endl(CONTENTS, '  data_phase_combination.sample();')
    contents_append_tab_endl(CONTENTS, 'end')
    contents_append_endl(CONTENTS, 'endfunction: sample_cov')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '/** get cov for cust_audio_data_phase_coverage_model */')
    contents_append_endl(CONTENTS, 'function real cust_audio_data_phase_coverage_model::get_cov();')
    contents_append_tab_endl(CONTENTS, 'get_cov = data_phase_combination.get_inst_coverage();')
    contents_append_endl(CONTENTS, 'endfunction: get_cov')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '`endif // CUST_AUDIO_DATA_PHASE_COVERAGE_MODEL__SV')
    
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_grapth(components_info):
    color_map = []
    G = nx.DiGraph()
    edge_labels = {}
    
    for component_info in components_info:
        G.add_node(int(component_info['NODE_ID']))
        if component_info['Type'] == 'Input_Node':
            color_map.append('green')
        elif component_info['Type'] == 'Output_Node':
            color_map.append('pink')
        else:
            color_map.append('gray')
        
    for component_info in components_info:
        node_id = int(component_info['NODE_ID'])
        for idx, select in enumerate(component_info['select']):
            if select != NULL_NODE_ID:
                G.add_edge(node_id, select, label = component_info['Inputs'][idx])
                edge_labels.update({(node_id, select): component_info['Inputs'][idx]})
            
            
    return G, color_map, edge_labels

def get_edges_from_path(path):
    edges = []
    for idx in range(len(path)-1):
        edges.append((path[idx], path[idx+1]))
    return edges

def path_list_to_string(path):
    expect_path_log = ""
    for node_id in path:
        expect_path_log += ' {}'.format(node_id)
    expect_path_log = expect_path_log[1:]
    
    return expect_path_log

def parse_register_file(input_file_name = 'all_register_dump_for_dv.txt', input_dir = INPUT_DIR):
    all_register_info = {}
    input_file_name = file_with_dir_path(input_file_name, input_dir)
    
    try:
        all_register_dump_for_dv_file = open(input_file_name, 'r')
    except:
        ERROR_MESSAGE('all_register_dump_for_dv.txt not found')
        control = input("Please prepare all_register_dump_for_dv.txt in the input directory, and press C(continue)")
        if control.lower() != 'c':
            return all_register_info
    
    for idx, line in enumerate(all_register_dump_for_dv_file.readlines()):
        if idx != 0:
            split_result = re.split('\s+', line)
            Signal = split_result[0]
            Addr   = split_result[1]
            Bits_Indices = split_result[2]
            
            delta = -1
            if Bits_Indices.find(':') != -1:
                start = int(Bits_Indices[:Bits_Indices.find(':')])
                end   = int(Bits_Indices[Bits_Indices.find(':')+1:])
                delta = int(abs(end-start))+1
            else:
                delta = 1
            
            all_register_info.update({ Signal: {'Addr': Addr, 'Bits_Indices': Bits_Indices, 'Length': delta}})
    
    return all_register_info

def set_value_reverse(set_value, IS_REVERSE):
    return_value = ''
    
    for value in set_value:
        if value == '1':
            if IS_REVERSE:
                return_value += '0'
            else:
                return_value += '1'
        else:
            if IS_REVERSE:
                return_value += '1'
            else:
                return_value += '0'
                
    return return_value

def add_signal_settings(CONTENTS, select_signal, set_value, signal_setting, node_id = -1, select_node_id = -1, IS_DEFAULT_SETTING = False):
    signal_setting = signal_setting.replace('set_value', set_value)
    
    if signal_setting and select_signal:
        if not IS_DEFAULT_SETTING:
            information = 'signal: {}, write value: {}, node[{}]({}) -> node[{}]({}) through {}'.format(select_signal, set_value, node_id, components_info[node_id]['Type'], select_node_id, components_info[select_node_id]['Type'], components_info[select_node_id]['Outputs'][0])
        else:
            information = 'signal: {}, write default value: {}'.format(select_signal, set_value)
            
        contents_append_tab_tab_endl(CONTENTS, '`uvm_info("audio_data_path_reg_config", "{}", UVM_LOW)'.format(information))
        
        for line in signal_setting.splitlines():
            contents_append_tab_tab_endl(CONTENTS, line)
            
        contents_append_endl(CONTENTS, '')
        
    return CONTENTS

def get_signal_default_settings(CONTENTS, node, SIGNAL_SETTINGS):

    select_signal = ''
    set_value = ''  
    signal_setting = ''
    

    ## 多個 selects -> 意味著 # of selects = # of inputs 且 selects 都為 1 bit
    if len(node['Selects']) > 1:

        for idx, select_signal in enumerate(node['Selects']):
            if select_signal != SELECT_SIGNAL_RESERVED:
                signal_setting = SIGNAL_SETTINGS[select_signal]['cmd_combo']
                set_value = set_value_reverse('0', node['IS_REVERSE'])
                signal_setting = signal_setting.replace('set_value', set_value)
                CONTENTS = add_signal_settings(CONTENTS, select_signal, set_value, signal_setting, IS_DEFAULT_SETTING = True)

    ## 單一 select
    ## case 1: 多 bits select (mul inputs)
    ## case 2:  1 bit  select (2   inputs)
    elif len(node['Selects']) == 1:
        select_signal = str(node['Selects'][0])
        if select_signal != SELECT_SIGNAL_RESERVED:
            signal_setting = SIGNAL_SETTINGS[select_signal]['cmd_combo']
            
            ## 多 bits select (mul inputs)
            if SIGNAL_SETTINGS[select_signal]['length'] > 1:
                set_value = ''
                bit_num = SIGNAL_SETTINGS[select_signal]['length']
                for i in range( bit_num - len(set_value)):
                    set_value += '0'
                set_value = set_value

                set_value = set_value_reverse( set_value, node['IS_REVERSE'])
                
            ## 1 bit  select
            else:
                set_value = set_value_reverse('0', node['IS_REVERSE'])
                
            CONTENTS = add_signal_settings(CONTENTS, select_signal, set_value, signal_setting, IS_DEFAULT_SETTING = True)
                
    return CONTENTS

def get_signal_setting(node_id, select_node_id, idx, components_info, SIGNAL_SETTINGS):
    node = components_info[node_id]
    select_node = components_info[select_node_id]
    
    select_signal = ''
    set_value = ''
    signal_setting = ''

    ## 多個 selects -> 意味著 # of selects = # of inputs 且 selects 都為 1 bit
    if len(node['Selects']) > 1:
        select_signal = str(node['Selects'][idx])
        if select_signal != SELECT_SIGNAL_RESERVED:
            signal_setting = SIGNAL_SETTINGS[select_signal]['cmd_combo']
            
            for i in range(len(node['Inputs'])):
                if i == idx:
                    set_value = set_value_reverse('1', node['IS_REVERSE'])
                    break
    
    ## 單一 selects
    ## case 1: 多 bits select (mul inputs)
    ## case 2:  1 bit  select (2   inputs)
    elif len(node['Selects']) == 1:
        select_signal = str(node['Selects'][0])
        if select_signal != SELECT_SIGNAL_RESERVED:
            signal_setting = SIGNAL_SETTINGS[select_signal]['cmd_combo']
            
            ## 多 bits select (mul inputs)
            if SIGNAL_SETTINGS[select_signal]['length'] > 1:
                for i in range(len(node['Inputs'])):
                    if idx == i:
                        dec_val = i
                        break
                        
                bit_num = SIGNAL_SETTINGS[select_signal]['length']
                set_value = bin(dec_val).replace('0b', '')
                padding_zero = ''
                for i in range( bit_num - len(set_value)):
                    padding_zero += '0'
                set_value = padding_zero + set_value

                set_value = set_value_reverse( set_value, node['IS_REVERSE'])
                
            ## 1 bit  select
            else:
                if idx == 1:
                    set_value = set_value_reverse('1', node['IS_REVERSE'])
                else:
                    set_value = set_value_reverse('0', node['IS_REVERSE'])
                
    return select_signal, set_value, signal_setting

def parse_in_and_out_info(info_files, input_dir):
    VIP_DEFINE_SETTINGS = {}
    VIP_ENABLE_SETTINGS = {}
    
    for info_file in info_files:
        info_file = file_with_dir_path(info_file, input_dir)
        wb = load_workbook(info_file)
        ws = wb.active
        col_names = []
        for cell in ws[1]:
            cell_value = cell.value
            if cell_value.replace(' ', '') != '':
                col_names.append(cell_value)
        col_names = col_names[1:]
        
        for row_idx, row in enumerate(ws):
            if row_idx != 0:
                VIP_DEFINE_SETTING = ''
                VIP_ENABLE_SETTING = ''
                for cell_idx, cell in enumerate(row[1:]):
                    if cell.value and cell.value == 1:
                        if col_names[cell_idx].find('pattern_define') != -1:
                            VIP_DEFINE_SETTING += (col_names[cell_idx] + '\n')
                            VIP_DEFINE_SETTINGS.update({ row[0].value : VIP_DEFINE_SETTING })
                        else:
                            VIP_ENABLE_SETTING += (col_names[cell_idx] + '\n')
                            VIP_ENABLE_SETTINGS.update({ row[0].value : VIP_ENABLE_SETTING })
    
    return VIP_DEFINE_SETTINGS, VIP_ENABLE_SETTINGS

def parse_reg_info(reg_info_file, input_dir):
    SIGNAL_SETTINGS = {}
    
    reg_info_file = file_with_dir_path(reg_info_file, input_dir)
    wb = load_workbook(reg_info_file)
    ws = wb.active
    #read_cmd_reg = REG_READ_WRITE_VARIABLE_NAME
    col_names = []
    for cell in ws[1]:
        cell_value = cell.value
        if cell_value.replace(' ', '') != '':
            col_names.append(cell_value)
            
    for row_idx, row in enumerate(ws):
        if row_idx != 0 and row[0].value and row[1].value and row[2].value and row[3].value and row[4].value:
            signal    = row[0].value
            address   = row[1].value
            bits      = row[2].value
            read_cmd  = row[3].value
            write_cmd = row[4].value
            cmd_combo = ''
            
            address = "{}'h{}".format(len(address)*4, address)
            #cmd_combo  = read_cmd + ', {}, {});\n'.format(address, read_cmd_reg)
            cmd_combo  = read_cmd.replace(REG_READ_WRITE_CMD_ADDR, address).replace(REG_READ_WRITE_VARIABLE, REG_READ_WRITE_VARIABLE_NAME) + '\n'
            cmd_combo += REG_READ_WRITE_VARIABLE_NAME + '[{}]'.format(bits)
            
            delta = -1
            if bits.find(':') != -1:
                start = int(bits[:bits.find(':')])
                end   = int(bits[bits.find(':')+1:])
                delta = int(abs(end-start))+1

                cmd_combo += " = {}'b".format(delta)
                cmd_combo += 'set_value;\n'
            else:
                cmd_combo += " = 1'b"
                cmd_combo += 'set_value;\n'
                delta = 1
                
            cmd_combo += write_cmd.replace(REG_READ_WRITE_CMD_ADDR, address).replace(REG_READ_WRITE_VARIABLE, REG_READ_WRITE_VARIABLE_NAME) + '\n'
            SIGNAL_SETTINGS.update( {signal : {'cmd_combo': cmd_combo, 'length': delta}} )
            
    return SIGNAL_SETTINGS

def cnt_SRC_in_path(path, components_info):
    SRC_NAME = 'SRC'
    SRC_cnt = 0.0
    
    for node in path:
        if components_info[int(node)]['Type'] == SRC_NAME:
            SRC_cnt += 1.0
    
    return math.ceil(SRC_cnt / 2.0)

def pattern_auto_gen(pattern_path, components_info, combination_idx, in_and_out_info_files = ['in_info.xlsx', 'out_info.xlsx'], reg_info_file = 'reg_info.xlsx', output_dir = PATTERN_DIR, input_dir = INPUT_DIR):
    VIP_DEFINE_SETTINGS, VIP_ENABLE_SETTINGS = parse_in_and_out_info(in_and_out_info_files, input_dir)
    SIGNAL_SETTINGS = parse_reg_info(reg_info_file, input_dir)

    if(os.path.isdir(output_dir)):
        shutil.rmtree(output_dir)
        
    for path_idx, stereo_path in enumerate(pattern_path):
        CONTENTS = []
        contents_append_endl(CONTENTS, SIGNATURE)
        
        class_name = 'audio_data_phase_auto_gen_{}'.format(path_idx)
        path_0 = stereo_path[0]
        path_1 = stereo_path[1]
        input_0  = int(path_0[-1])
        input_1  = int(path_1[-1])
        output_0 = int(path_0[0])
        output_1 = int(path_1[0])
        input_signal  = components_info[input_0]['Outputs'][0]
        output_signal = components_info[output_0]['Inputs'][0]
        
        ## class init
        if input_signal in VIP_DEFINE_SETTINGS:
            contents_append_endl(CONTENTS, VIP_DEFINE_SETTINGS[input_signal])
        if output_signal in VIP_DEFINE_SETTINGS:
            contents_append_endl(CONTENTS, VIP_DEFINE_SETTINGS[output_signal])

        # contents_append_endl(CONTENTS, 'import "DPI-C" context function int uvm_hdl_force(string path, uvm_hdl_data_t value);')
        # contents_append_endl(CONTENTS, 'import "DPI-C" context function int uvm_hdl_release(string path);')
            
        contents_append_endl(CONTENTS, 'class {} extends audio_data_phase_golden;'.format(class_name))
        contents_append_tab_endl(CONTENTS, '`uvm_component_utils({})'.format(class_name))
        
        contents_append_endl(CONTENTS, '')
        
        ## new function
        contents_append_tab_endl(CONTENTS, 'function new(string name, uvm_component parent);')
        contents_append_tab_tab_endl(CONTENTS, 'super.new(name, parent);')
        contents_append_tab_endl(CONTENTS, 'endfunction')
        
        contents_append_endl(CONTENTS, '')

        ## build phase
        contents_append_tab_endl(CONTENTS, 'virtual function void build_phase(uvm_phase phase);')
        contents_append_tab_tab_endl(CONTENTS, 'super.build_phase(phase);')
        if input_signal in VIP_ENABLE_SETTINGS:
            for line in VIP_ENABLE_SETTINGS[input_signal].splitlines():
                contents_append_tab_tab_endl(CONTENTS, line)
                
        if output_signal in VIP_ENABLE_SETTINGS:
            for line in VIP_ENABLE_SETTINGS[output_signal].splitlines():
                contents_append_tab_tab_endl(CONTENTS, line)
                
            # audo gen channel_enable
        for idx, output_node in enumerate(get_nodes_by_type(components_info, 'Output_Node')):
            if output_node['NODE_ID'] == components_info[output_0]['NODE_ID']:
                contents_append_tab_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].channel_enable = 1;' % idx)

                # amp adjustment
                SRC_cnt   = cnt_SRC_in_path(path_0, components_info)
                SRC_decay = 1.5
                contents_append_tab_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].amp_threshold -= %d;' % (idx, SRC_cnt * SRC_decay))
            
            if output_node['NODE_ID'] == components_info[output_1]['NODE_ID']:
                contents_append_tab_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].channel_enable = 1;' % idx)

                # amp adjustment
                SRC_cnt   = cnt_SRC_in_path(path_1, components_info)
                SRC_decay = 1.5
                contents_append_tab_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].amp_threshold -= %d;' % (idx, SRC_cnt * SRC_decay))

        # expect log
        expect_path_log = path_list_to_string(path_0)
        contents_append_tab_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].expect_path_log_0 = "{}";'.format(expect_path_log))
        expect_path_log = path_list_to_string(path_1)
        contents_append_tab_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].expect_path_log_1 = "{}";'.format(expect_path_log))

        # for data phase, combination index
        contents_append_tab_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].combination_idx = new[{}];'.format(len(combination_idx[path_idx])))
        combination_idx_arr_str = "sys_cfg.audio_data_path_cfg[0].combination_idx = '{" + str(combination_idx[path_idx][0])
        for i in range(len(combination_idx[path_idx])):
            if i != 0:
                combination_idx_arr_str += ', ' + str(combination_idx[path_idx][i])
        combination_idx_arr_str += '};'
        contents_append_tab_tab_endl(CONTENTS, combination_idx_arr_str)
        
        
        
        contents_append_tab_endl(CONTENTS, 'endfunction')
        
        contents_append_endl(CONTENTS, '')
        
        ## reset_phase
        contents_append_tab_endl(CONTENTS, 'virtual task reset_phase(uvm_phase phase);')
        contents_append_tab_tab_endl(CONTENTS, 'super.reset_phase(phase);')
        contents_append_tab_endl(CONTENTS, 'endtask')
        
        contents_append_endl(CONTENTS, '')
        
        ## audio_data_path_default_reg_config
        contents_append_tab_endl(CONTENTS, 'virtual task audio_data_path_default_reg_config();')
        contents_append_endl(CONTENTS, '')
        for component_info in components_info:
            CONTENTS = get_signal_default_settings(CONTENTS, component_info, SIGNAL_SETTINGS)
                
        contents_append_tab_endl(CONTENTS, 'endtask')
        contents_append_endl(CONTENTS, '')
        
        ## audio_data_path_reg_config
        contents_append_tab_endl(CONTENTS, 'virtual task audio_data_path_reg_config();')
        contents_append_endl(CONTENTS, '')
        
        edges  = get_edges_from_path(path_0)
        for edge in edges:
            node   = components_info[edge[0]]
            select = edge[1]

            for idx, select_node in enumerate(node['select']):
                if select == select_node:
                    select_signal, set_value, signal_setting = get_signal_setting(edge[0], edge[1], idx, components_info, SIGNAL_SETTINGS)
                    CONTENTS = add_signal_settings(CONTENTS, select_signal, set_value, signal_setting, edge[0], edge[1])
                    break
        
        edges  = get_edges_from_path(path_1)
        for edge in edges:
            node   = components_info[edge[0]]
            select = edge[1]

            for idx, select_node in enumerate(node['select']):
                if select == select_node:
                    select_signal, set_value, signal_setting = get_signal_setting(edge[0], edge[1], idx, components_info, SIGNAL_SETTINGS)
                    CONTENTS = add_signal_settings(CONTENTS, select_signal, set_value, signal_setting, edge[0], edge[1])
                    break
                        
        contents_append_tab_endl(CONTENTS, 'endtask')
        
        contents_append_endl(CONTENTS, '')
        
        ## main_phase
        contents_append_tab_endl(CONTENTS, 'virtual task main_phase(uvm_phase phase);')
        path_string = ''
        for idx, node_id in enumerate(reversed(path_0)):
            path_string += '{}'.format(node_id)
            if idx == 0:
                path_string += '(Input) '
            elif idx == (len(path_0)-1) :
                path_string += '(Output) '
            else:
                path_string += ' '

        contents_append_tab_tab_endl(CONTENTS, '`uvm_info("audio_data_path_pattern", "Path[{}]: {}", UVM_LOW)'.format(path_idx, path_string))
        for idx, node_id in enumerate(path_0):
            node = components_info[node_id]
            temp_content  = '`uvm_info("audio_data_path_pattern", "Node['
            temp_content += '%3s' % node_id
            temp_content += ']('
            temp_content += '%35s' % node['Type']
            
            if node['Type'] == 'Output_Node':
                sig_info = node['Inputs'][0]
                temp_content += '),  input signal: '
                temp_content += '%35s' % sig_info
                temp_content += '", UVM_LOW)'
                contents_append_tab_tab_endl(CONTENTS, temp_content)
            else:
                sig_info = node['Outputs'][0]
                temp_content += '), output signal: '
                temp_content += '%35s' % sig_info
                temp_content += '", UVM_LOW)'
                contents_append_tab_tab_endl(CONTENTS, temp_content)

        contents_append_endl(CONTENTS, '')
        
        path_string = ''
        for idx, node_id in enumerate(reversed(path_1)):
            path_string += '{}'.format(node_id)
            if idx == 0:
                path_string += '(Input) '
            elif idx == (len(path_1)-1) :
                path_string += '(Output) '
            else:
                path_string += ' '

        contents_append_tab_tab_endl(CONTENTS, '`uvm_info("audio_data_path_pattern", "Path[{}]: {}", UVM_LOW)'.format(path_idx, path_string))
        for idx, node_id in enumerate(path_1):
            node = components_info[node_id]
            temp_content  = '`uvm_info("audio_data_path_pattern", "Node['
            temp_content += '%3s' % node_id
            temp_content += ']('
            temp_content += '%35s' % node['Type']
            
            if node['Type'] == 'Output_Node':
                sig_info = node['Inputs'][0]
                temp_content += '),  input signal: '
                temp_content += '%35s' % sig_info
                temp_content += '", UVM_LOW)'
                contents_append_tab_tab_endl(CONTENTS, temp_content)
            else:
                sig_info = node['Outputs'][0]
                temp_content += '), output signal: '
                temp_content += '%35s' % sig_info
                temp_content += '", UVM_LOW)'
                contents_append_tab_tab_endl(CONTENTS, temp_content)
            
        contents_append_tab_tab_endl(CONTENTS, 'super.main_phase(phase);')
        contents_append_tab_endl(CONTENTS, 'endtask')
        
        contents_append_endl(CONTENTS, '')
        
        ## end class
        contents_append_endl(CONTENTS, 'endclass: {}'.format(class_name))
        
        output_file_name = class_name + '.sv'
        write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def read_components_info(input_file_name = 'components_info.txt', input_dir = DATA_DIR):
    components_info = []
    input_file_name = file_with_dir_path(input_file_name, input_dir)

    with open(input_file_name) as txt_file:
        for line in txt_file.readlines():
            component_info = json.loads(line[line.find('{'):-1])
            # component_info.pop('select', None)
            components_info.append(component_info)

    return components_info

def output_components_info(components_info, output_file_name = 'components_info.txt', output_dir = DATA_DIR):
    CONTENTS = []
    for idx, component_info in enumerate(components_info):
        contents_append_endl(CONTENTS, 'Component[%3s]: ' % str(idx) + json.dumps(component_info))
        
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_input_output_stereo_dict():
    input_output_stereo_dict = dict()

    with open('input/stereo_table.txt', 'r') as f:
        lines = f.readlines()

    for line in lines:
        input_output_stereo_dict[line.split(' ')[1]] = line.split(' ')[2].replace('\n', '')
    # print(input_output_stereo_dict)
    # print(len(input_output_stereo_dict))
    
    return input_output_stereo_dict

def gen_mux_stereo_dict(components_info):
    mux_sel_signal_arr = []
    for i in range(len(components_info)):
        if components_info[i]['Type'].find('MUX') != -1:
            #print(components_info[i]['Type'], components_info[i]['Outputs'], components_info[i]['Selects'], components_info[i]['Inputs'], components_info[i]['select'])
            #print(components_info[i]['Selects'][0])
            mux_sel_signal_arr.append(components_info[i]['Selects'][0])

    print('total mux select signal num:', len(mux_sel_signal_arr))
    print('non repeat mux select signal num:', len(set(mux_sel_signal_arr)))

    seen = set()
    dupes = set()
    for x in mux_sel_signal_arr:
        if x in seen:
            dupes.add(x)
        else:
            seen.add(x)

    uniq = list(set(mux_sel_signal_arr) - dupes)
    dupes = list(dupes)
    
    print('----------------------------------------------------')
    print('duplicated:', len(dupes))
    print(dupes)
    print('----------------------------------------------------')
    print('uniq:', len(uniq))
    print(uniq)

    print('----------------------------------------------------')

    if len(dupes) + len(uniq) != len(set(mux_sel_signal_arr)):
        print('select signal count error!!!!!!!!!!!!')
        exit()

    if len(dupes)*2 + len(uniq) == len(mux_sel_signal_arr):
        print('PASS! Each duplicated signal has two')
    else:
        print('Warning! Some duplicated signals number is not two')
        
        # for i in range(len(dupes)):
        #     cou = 0
        #     for j in range(len(mux_sel_signal_arr)):
        #         if dupes[i] == mux_sel_signal_arr[j]:
        #             cou += 1
        #     print(dupes[i], cou)
        
        # print('----------------------------------------------------')

        # for i in range(len(uniq)):
        #     cou = 0
        #     for j in range(len(mux_sel_signal_arr)):
        #         if uniq[i] == mux_sel_signal_arr[j]:
        #             cou += 1
        #     print(uniq[i], cou)
        # exit()

    mux_uniq_stereo_dict = dict()                 # two stereo mux select signals are different (unique), ex: XXX_sel_r, XXX_sel_l
    mux_duplicate_stereo_dict = dict()            # two stereo mux select signals are same (duplicate), ex: XXX_sel, XXX_sel
    mux_non_pair_stereo_dict = dict()             # cannot find stereo pairs
    find_pairs_names = []

    for item in combinations(uniq, 2):
        len_0 = len(item[0])
        len_1 = len(item[1])
        if len_0 == len_1:
            S = SequenceMatcher(None, item[0], item[1])
            r = S.ratio()
            # print(len_0 - (r*(len_0*2)/2))
            if len_0 - (r*(len_0*2)/2) <= 1.0:    # item[0] 跟 item[1] 字串只差一個字母
                #print(item)
                diff_idx = S.get_matching_blocks()[0][2] # 相差字母的位置
                #print(diff_idx)
                if item[0][diff_idx] == 'l':
                    if item[1][diff_idx] == 'r':
                        mux_uniq_stereo_dict[item[0]] = item[1]
                        mux_uniq_stereo_dict[item[1]] = item[0]
                        find_pairs_names.append(item[0])
                        find_pairs_names.append(item[1])
                        
                if item[0][diff_idx] == 'r':
                    if item[1][diff_idx] == 'l':
                        mux_uniq_stereo_dict[item[0]] = item[1]
                        mux_uniq_stereo_dict[item[1]] = item[0]
                        find_pairs_names.append(item[0])
                        find_pairs_names.append(item[1])
    
    #print(len(mux_uniq_stereo_dict), mux_uniq_stereo_dict)

    print('----------------------------------------------------')
    non_pair_uniq = list(set(uniq) - set(find_pairs_names))
    print('沒有成對的 uniq :', non_pair_uniq)

    for item in dupes:
        mux_duplicate_stereo_dict[item] = item

    for item in non_pair_uniq:
        mux_non_pair_stereo_dict[item] = item

    print('----------------------------------------------------')
    print('total mux stereo dict num:', len(mux_uniq_stereo_dict)+len(mux_duplicate_stereo_dict)+len(mux_non_pair_stereo_dict))
    print('----------------------------------------------------')
    print('uniq dict')
    print(mux_uniq_stereo_dict)
    print('----------------------------------------------------')
    print('duplicate dict')
    print(mux_duplicate_stereo_dict)
    print('----------------------------------------------------')
    print('non pair dict')
    print(mux_non_pair_stereo_dict)

    print('----------------------------------------------------')
    if len(mux_uniq_stereo_dict)+len(mux_duplicate_stereo_dict)+len(mux_non_pair_stereo_dict) != len(set(mux_sel_signal_arr)):
        print('some select signals didnt find stereo pairs')
    else:
        print('DONE')

    return mux_uniq_stereo_dict, mux_duplicate_stereo_dict, mux_non_pair_stereo_dict

def only_direct_edge(G, node_0, node_1):
    if  (node_0, node_1) in G.edges and (node_1, node_0) not in G.edges:
        return True
    else:
        return False

def find_stereo_path(G, components_info, input_node_id, input_output_stereo_dict, mux_uniq_stereo_dict, mux_duplicate_stereo_dict, mux_non_pair_stereo_dict, path_0):
    # give one path(path_0) -> find another stereo path(path_1)
    path_1 = []                # target
    path_1_len = 0             # path_1 now path length
    success_flag = True
    path_1_output_name = input_output_stereo_dict[components_info[path_0[0]]['Inputs'][0]]   # find path_1 output node name
    for i in range(len(components_info)):
        if components_info[i]['Inputs'] == [path_1_output_name]:
            path_1.append(i)                                                                 # append path_1 output node index
            path_1_len += 1
            break

    while(path_1_len != len(path_0)):
        try:
            path_0_next_sel_idx = components_info[path_0[path_1_len-1]]['select'].index(path_0[path_1_len])
            path_1_next_sel_id  = components_info[path_1[path_1_len-1]]['select'][path_0_next_sel_idx]
            if path_1_next_sel_id == -1:                                                     # -1 means this select is null
                success_flag = False
                break
            
            if components_info[path_1[path_1_len-1]]['Type'].find('MUX') != -1:
                path_0_sel_signal = components_info[path_0[path_1_len-1]]['Selects'][0]
                path_1_sel_signal = components_info[path_1[path_1_len-1]]['Selects'][0]

                if path_0_sel_signal in mux_uniq_stereo_dict:
                    if mux_uniq_stereo_dict[path_0_sel_signal] != path_1_sel_signal:
                        success_flag = False
                        break

                    if path_0_next_sel_idx%2 == 0:
                        path_0_next_sel_idx += 1
                    else:
                        path_0_next_sel_idx -= 1
                    
                    path_1_next_sel_id  = components_info[path_1[path_1_len-1]]['select'][path_0_next_sel_idx]
                    if path_1_next_sel_id == -1:
                        success_flag = False
                        break
                    
                elif path_0_sel_signal in mux_duplicate_stereo_dict:
                    if mux_duplicate_stereo_dict[path_0_sel_signal] != path_1_sel_signal:
                        success_flag = False
                        break
                elif path_0_sel_signal in mux_non_pair_stereo_dict:
                    if mux_non_pair_stereo_dict[path_0_sel_signal] != path_1_sel_signal:
                        success_flag = False
                        break

            path_1.append(path_1_next_sel_id)
            path_1_len += 1
        except:
            success_flag = False
            #print('idx error')
            break


    # 檢查兩個 stereo paths 的所有 node 數量 是否相同
    if success_flag:
        if len(set(path_0+path_1)) != len(path_0+path_1):
            success_flag = False
            # for item in (set(path_0) & set(path_1)):       # set(path_0) & set(path_1) -> path_0 跟 path_1 重複的點的集合
            #     if item not in input_node_id:
            #         print('Stereo path found repeated node not in input_node list !!!!')
            #         success_flag = False

    if success_flag:
        path_1_expect_input_node_name = input_output_stereo_dict[components_info[path_0[-1]]['Outputs'][0]]          # find path_1 input node name
        if path_1_expect_input_node_name != components_info[path_1[-1]]['Outputs'][0]:                               # 若找到的另一條stereo_path的input node跟原本的input node沒有成對
            success_flag = False
            print('input node diff,', path_1_expect_input_node_name, components_info[path_1[-1]]['Outputs'][0])
            print(path_0)
            print(path_1)
            exit()

    #print(path_1)
    return success_flag, path_1

def check_per_path(G, node_0, node_1, node_2, node_3=None):
    success_flag = True
    arr = []
    arr.append(node_0)
    arr.append(node_1)
    arr.append(node_2)
    if node_3 != None:
        arr.append(node_3)

    for i in range(len(arr)-1):
        for j in range(i+1, len(arr)):
            if nx.has_path(G, arr[i], arr[j]):
                pass
            else:
                success_flag = False
                break
        if success_flag == False:
            break
    return success_flag

def permutation_find_path(G, node_0, node_1, node_2, node_3=None):
    if node_3==None:
        PATH_A = [node_0, node_1]
        PATH_B = [node_1, node_2]
        path_list = [PATH_A, PATH_B]
        per_list = permutations(path_list)
    else:
        PATH_A = [node_0, node_1]
        PATH_B = [node_1, node_2]
        PATH_C = [node_2, node_3]
        path_list = [PATH_A, PATH_B, PATH_C]
        per_list = permutations(path_list)
    
    mix = []                  # 排列組合後 組合而成的path
    success_flag = False

    per = permutations(range(2))
    per_table_sort_2 = np.argsort([i for i in per])
    per = permutations(range(3))
    per_table_sort_3 = np.argsort([i for i in per])

    for i, item in enumerate(per_list):
        #print(item)
        #print(item[per_table_sort[i][0]], item[per_table_sort[i][1]], item[per_table_sort[i][2]])

        per_pick_nodes = set([y for x in item for y in x])

        remove_node = per_pick_nodes - set([x for x in item[0]])
        G_tmp = G.copy()
        for node in remove_node:
            G_tmp.remove_node(node)
        if nx.has_path(G_tmp, item[0][0], item[0][1]):
            all_path_a = nx.all_shortest_paths(G_tmp, item[0][0], item[0][1])
        else:
            continue

        for a in all_path_a:
            remove_node = (per_pick_nodes | set(a)) - set([x for x in item[1]])
            G_tmp = G.copy()
            for node in remove_node:
                G_tmp.remove_node(node)
            if nx.has_path(G_tmp, item[1][0], item[1][1]):
                all_path_b = nx.all_shortest_paths(G_tmp, item[1][0], item[1][1])
            else:
                continue

            for b in all_path_b:
                if node_3==None:
                    combine_path = [a, b]
                    mix = combine_path[per_table_sort_2[i][0]] + combine_path[per_table_sort_2[i][1]][1:]
                    success_flag = True
                    break
                remove_node = (per_pick_nodes | set(a) | set(b)) - set([x for x in item[2]])
                G_tmp = G.copy()
                for node in remove_node:
                    G_tmp.remove_node(node)
                if nx.has_path(G_tmp, item[2][0], item[2][1]):
                    c = nx.shortest_path(G_tmp, item[2][0], item[2][1])
                    combine_path = [a, b, c]
                    mix = combine_path[per_table_sort_3[i][0]] + combine_path[per_table_sort_3[i][1]][1:] + combine_path[per_table_sort_3[i][2]][1:]
                    success_flag = True
                    break
            if success_flag == True:
                    break
    
    return success_flag, mix

def gen_find_path(pn2_has_edge):
    find_path_flag = False
    find_path = []
    not_found_path_node = []
    illegal_stereo_path_node = []
    
    for i in range(len(pn2_has_edge)):
        find_path_flag = False
        print('{:>4d}->{:<4d}: '.format(pn2_has_edge[i][0], pn2_has_edge[i][1]), end='')
        start_time = time.time()
        if components_info[pn2_has_edge[i][1]]['Type'] == 'Input_Node':
            for output_node in output_node_id:
                # if output_node <= 321:
                #     continue
                if check_per_path(G, output_node, pn2_has_edge[i][0], pn2_has_edge[i][1]):
                    tmp_a = nx.shortest_path(G, output_node, pn2_has_edge[i][0])
                    tmp_b = nx.shortest_path(G, pn2_has_edge[i][0], pn2_has_edge[i][1])
                    tmp_mix = tmp_a + tmp_b[1:]
                    
                    if len(set(tmp_mix)) != len(tmp_mix):
                        #print(tmp_mix)
                        # print('GG')
                        find_path_flag, tmp_mix = permutation_find_path(G, output_node, pn2_has_edge[i][0], pn2_has_edge[i][1])
                        if find_path_flag == False:
                            continue

                    find_path_flag, tmp_mix_stereo = find_stereo_path(G, components_info, input_node_id, input_output_stereo_dict, mux_uniq_stereo_dict, mux_duplicate_stereo_dict, mux_non_pair_stereo_dict, tmp_mix)
                    if find_path_flag == False:
                        if [pn2_has_edge[i][0], pn2_has_edge[i][1]] not in illegal_stereo_path_node:
                            illegal_stereo_path_node.append([pn2_has_edge[i][0], pn2_has_edge[i][1]])
                        continue
                    
                    # print()
                    # print(tmp_mix)
                    # print(tmp_mix_stereo)
                    find_path.append([tmp_mix, tmp_mix_stereo])
                    find_path_flag = True
                    break
        else:
            for output_node in output_node_id:
                # if output_node <= 321:
                #     continue
                for input_node in input_node_id:
                    #print('Try ', output_node, input_node)
                    if check_per_path(G, output_node, pn2_has_edge[i][0], pn2_has_edge[i][1], input_node):
                        tmp_a = nx.shortest_path(G, output_node, pn2_has_edge[i][0])
                        tmp_b = nx.shortest_path(G, pn2_has_edge[i][0], pn2_has_edge[i][1])
                        tmp_c = nx.shortest_path(G, pn2_has_edge[i][1], input_node)
                        tmp_mix = tmp_a + tmp_b[1:] + tmp_c[1:]

                        if len(set(tmp_mix)) != len(tmp_mix):
                            #print(tmp_mix)
                            #print('QQ')
                            find_path_flag, tmp_mix = permutation_find_path(G, output_node, pn2_has_edge[i][0], pn2_has_edge[i][1], input_node)
                            if find_path_flag == False:
                                continue
                        
                        find_path_flag, tmp_mix_stereo = find_stereo_path(G, components_info, input_node_id, input_output_stereo_dict, mux_uniq_stereo_dict, mux_duplicate_stereo_dict, mux_non_pair_stereo_dict, tmp_mix)
                        if find_path_flag == False:
                            if [pn2_has_edge[i][0], pn2_has_edge[i][1]] not in illegal_stereo_path_node:
                                illegal_stereo_path_node.append([pn2_has_edge[i][0], pn2_has_edge[i][1]])
                            continue
                        
                        # print()
                        # print(tmp_mix)
                        # print(tmp_mix_stereo)
                        find_path.append([tmp_mix, tmp_mix_stereo])
                        find_path_flag = True
                        break
                if find_path_flag == True:
                    break

        if find_path_flag == False:
            print('FAIL !!! some shortest path of pn2_has_edge not found')
            not_found_path_node.append([pn2_has_edge[i][0], pn2_has_edge[i][1]])
            #exit()
        else:
            print('FIND_PATH')
        
        # if time.time() - start_time > 5:
        #     print(time.time() - start_time)

    return find_path, not_found_path_node, illegal_stereo_path_node

def greedy_pick_path(find_path, pn2_has_edge):
    greedy_choose_path = []          # 用greedy演算法選到的path
    max_match_info = [0, 0, []]      # [max_match_num, find_path idx, covered_pairs_idx array]
    useless_path = []                # 存放在 find_path 中沒有用的 path idx (len(covered_pairs_idx)==0)
    last_max_match_num = -1          # 上一次找到的 max_match_num 值
    start_idx = 0                    # 從 start_idx 開始找 find_path for loop
    DEF_MAX_VAULE = 99999999         # 
    end_idx = DEF_MAX_VAULE

    remain_path = find_path.copy()
    uncover_pairs = pn2_has_edge.copy()
    
    start_time = time.time()

    while(True):
        if end_idx == DEF_MAX_VAULE:
            max_match_info = [0, 0, []]
            useless_path = []
        print('remain path :', len(remain_path), ',   uncovered combination pairs :', len(uncover_pairs))
        for path_idx in range(len(remain_path)):
            if path_idx < start_idx or path_idx > end_idx:      # start_idx >= path_idx >= end_idx  not continue
                continue
            covered_pairs_idx = []
            for pair_idx in range(len(uncover_pairs)):
                mono_0_path = remain_path[path_idx][0]
                for mono_0_pre in range(len(mono_0_path)):
                    if uncover_pairs[pair_idx][0] == mono_0_path[mono_0_pre]:
                        for mono_0_post in range(mono_0_pre+1, len(mono_0_path)):
                            if uncover_pairs[pair_idx][1] == mono_0_path[mono_0_post]:
                                #print('mono_0 find pair :', pair_idx, uncover_pairs[pair_idx])
                                covered_pairs_idx.append(pair_idx)

                mono_1_path = remain_path[path_idx][1]
                for mono_1_pre in range(len(mono_1_path)):
                    if uncover_pairs[pair_idx][0] == mono_1_path[mono_1_pre]:
                        for mono_1_post in range(mono_1_pre+1, len(mono_1_path)):
                            if uncover_pairs[pair_idx][1] == mono_1_path[mono_1_post]:
                                #print('mono_1 find pair :', pair_idx, uncover_pairs[pair_idx])
                                covered_pairs_idx.append(pair_idx)
            
            if len(covered_pairs_idx) == 0:
                useless_path.append(path_idx)

            if max_match_info[0] < len(covered_pairs_idx):      # 如果當下找到的path能cover更多pn2組合
                max_match_info[0] = len(covered_pairs_idx)
                max_match_info[1] = path_idx
                max_match_info[2] = covered_pairs_idx
                if last_max_match_num == max_match_info[0]:     # 如果目前找到的 max_match_num 是上一次找到的 max_match_num(last_max_match_num) 就可以不用繼續找了，可直接把這個path存起來
                    break
            elif end_idx != DEF_MAX_VAULE and max_match_info[0] == len(covered_pairs_idx) and max_match_info[1] > path_idx:
                max_match_info[0] = len(covered_pairs_idx)
                max_match_info[1] = path_idx
                max_match_info[2] = covered_pairs_idx

            # print(covered_pairs_idx)
            # print(len(covered_pairs_idx))
        
        if max_match_info[0] == 0:       # cannot found uncover_pairs anymore
            if start_idx != 0:
                start_idx = 0
                continue
            else:
                break
        
        if last_max_match_num != max_match_info[0] and start_idx != 0:    # 代表這次尋找是從中間段開始搜尋，但已經找不到 last_max_match_num 的匹配數量了
            end_idx = start_idx
            start_idx = 0
            continue

        
        last_max_match_num = max_match_info[0]                 # update last_max_match_num value
        start_idx = max_match_info[1] - len(useless_path)
        if start_idx < 0:
            start_idx = 0
        end_idx = DEF_MAX_VAULE

        greedy_choose_path.append(remain_path[max_match_info[1]])
        useless_path.append(max_match_info[1])
        delete_tmp = sorted(useless_path, reverse=True)
        for item in delete_tmp:
            del remain_path[item]
        delete_tmp = max_match_info[2][::-1]
        for item in delete_tmp:
            del uncover_pairs[item]
        
        if len(uncover_pairs) == 0:       # covered every pair in uncover_pairs -> finish
            break
    
    print(time.time()-start_time)

    #print(uncover_pairs)
    #print(pn2_has_edge)
    
    return greedy_choose_path, uncover_pairs

def read_pickle(file_name):
    file_dir = data_phase_array_data_dir + file_name + '.pickle'
    if os.path.exists(file_dir):
        with open(file_dir, 'rb') as f:
            data = pickle.load(f)
        return data
    else:
        print(file_dir, 'not exist!!')
        exit()

def dump_pickle(file_name, data):
    global cust_audio_data_phase_coverage_model_changed
    file_dir = data_phase_array_data_dir + file_name + '.pickle'
    c = 'y'
    if os.path.exists(file_dir):
        with open(file_dir, 'rb') as f:
            old_data = pickle.load(f)
        if old_data != data:          # if data changed
            print(file_name + ' data changed...  backup old data!')
            src = file_dir
            des = data_phase_array_data_dir + 'backup/' + file_name + '.pickle'
            c = input('dump?')
            if c == 'y':
                shutil.move(src, des)
                if file_name == 'pn2_has_edge_covered':
                    cust_audio_data_phase_coverage_model_changed = True
    else:
        if file_name == 'pn2_has_edge_covered':
            cust_audio_data_phase_coverage_model_changed = True
    if c == 'y':
        with open(file_dir, 'wb') as f:
            pickle.dump(data, f)

def cov_self_check(input_dir=PATTERN_DIR):
    with open(ENV_DIR + '/cust_audio_data_phase_coverage_model.sv', 'r') as f:
        coverage = f.readlines()
    cov_arr = []
    for cov in coverage:
        if cov.find('bins C_') != -1:
            #print(cov)
            cov_list = cov.split('C_')[-1].split(' ')[0].split('_')
            cov_list_map = map(int, cov_list)
            cov_list_int = list(cov_list_map)
            cov_arr.append(cov_list_int)
    #print(cov_arr, len(cov_arr))
    print('coverage bins num:', len(cov_arr))
    cov_check = [0]*len(cov_arr)

    check_pass = True

    if(os.path.isdir(input_dir)):
        for dirPath, dirNames, fileNames in os.walk(input_dir):
            for f in fileNames:
                with open(os.path.join(dirPath, f), 'r') as ptr_f:
                    pattern = ptr_f.readlines()

                for item in pattern:
                    if item.find('expect_path_log_0') != -1:
                        path_0_str = item.split('"')[1].split(' ')
                        path_0_map = map(int, path_0_str)
                        path_0 = list(path_0_map)
                        #print(path_0)
                        
                    if item.find('expect_path_log_1') != -1:
                        path_1_str = item.split('"')[1].split(' ')
                        path_1_map = map(int, path_1_str)
                        path_1 = list(path_1_map)
                        #print(path_1)
                    
                    if item.find("combination_idx = '") != -1:
                        combination_idx_str = item.split('{')[-1].split('}')[0].replace(',', '').split(' ')
                        combination_idx_map = map(int, combination_idx_str)
                        combination_idx = list(combination_idx_map)
                        #print(combination_idx)

                for idx in combination_idx:
                    first_node = cov_arr[idx][0]
                    second_node = cov_arr[idx][1]
                    #print(cov_arr[idx])
                    
                    success_flag = False
                    for i in range(len(path_0)):
                        if path_0[i] == first_node:
                            for j in range(i, len(path_0)):
                                if path_0[j] == second_node:
                                    success_flag = True
                                    break
                    
                    if success_flag == False:
                        for a in range(len(path_1)):
                            if path_1[a] == first_node:
                                for b in range(a, len(path_1)):
                                    if path_1[b] == second_node:
                                        success_flag = True
                                        break
                    
                    if success_flag == False:
                        print('coverage ERROR!!!!')
                        print('filename:', f)
                        print('path 0:', path_0)
                        print('path 1:', path_1)
                        print('not found combination_idx:', idx)
                        exit()
                    else:
                        cov_check[idx] = 1
        
    #print(cov_check)
    for i in range(len(cov_check)):
        if cov_check[i] == 0:
            print('uncovered bins idx:', i)
            check_pass = False

    if check_pass:
        print('coverage self check PASS')
    else:
        print('coverage self check FAIL!!')



if __name__ == '__main__':
    ################################################################## main ##################################################################

    data_phase_array_data_dir = 'input/data_phase/'

    if(os.path.isdir('input/data_phase/') == False):
        os.mkdir('input/data_phase/')
    if(os.path.isdir('input/data_phase/backup/') == False):
        os.mkdir('input/data_phase/backup/')

    
    components_info = read_components_info()


    # find MIX_XXXX components
    mix_permuation_pair = []                              # these pairs are impossible to construct a stereo path
    mix_inputs_components = []                            # these components will add in data_phase_permutation_nodes array
    for i in range(len(components_info)):
        if components_info[i]['Type'].find('MIX') != -1:
            mix_num = len(components_info[i]['Inputs'])   # how many input nodes in this mix
            # print(mix_num)
            permutation_pair = []
            for mix_idx in range(mix_num):
                if components_info[i]['select'][mix_idx] != -1:
                    permutation_pair.append(components_info[i]['select'][mix_idx])
                    mix_inputs_components.append(components_info[i]['select'][mix_idx])
                # mix_input_name = components_info[i]['Inputs'][mix_idx]
                # for j in range(len(components_info)):
                #     if components_info[j]['Outputs'] == [mix_input_name]:
                #         permutation_pair.append(j)
                #         mix_inputs_components.append(j)
            if len(permutation_pair) >= 2:
                for item in permutations(permutation_pair, 2):
                    mix_permuation_pair.append(item)
    # print(mix_permuation_pair)
    # print(mix_inputs_components)
    # print(len(mix_inputs_components), len(set(mix_inputs_components)))

    if len(mix_inputs_components) != len(set(mix_inputs_components)):
        print('Warning!! mix inputs repeat')

    print('gen_grapth...')
    G, color_map, edge_labels = gen_grapth(components_info)
    print('done\n')


    input_node_id = []
    output_node_id = []
    for i in range(len(components_info)):
        if components_info[i]['Type'] == 'Input_Node':
            input_node_id.append(i)
        if components_info[i]['Type'] == 'Output_Node':
            output_node_id.append(i)

    input_output_stereo_dict = gen_input_output_stereo_dict()
    mux_uniq_stereo_dict, mux_duplicate_stereo_dict, mux_non_pair_stereo_dict = gen_mux_stereo_dict(components_info)

    data_phase_permutation_nodes = []
    for i in range(len(components_info)):
        if components_info[i]['Type'] == 'Block' or components_info[i]['Type'] == 'SRC' or components_info[i]['Type'] == 'Input_Node':
            # print(components_info[i]['Type'])
            data_phase_permutation_nodes.append(i)

    for item in mix_inputs_components:
        if item not in data_phase_permutation_nodes:
            data_phase_permutation_nodes.append(item)

    if len(data_phase_permutation_nodes) != len(set(data_phase_permutation_nodes)):
        print('Error!! data_phase_permutation_nodes repeat')
        exit()
    
    print('------------------------------------------------------------')
    print('permutation node', data_phase_permutation_nodes)
    print('permutation node numbers : ' + str(len(data_phase_permutation_nodes)))
    per_list = permutations(data_phase_permutation_nodes, 2)

    print('------------------------------------------------------------')
    print('check permutation pairs has edge...')
    
    if input('Read data from files?(y/n) ').lower() == 'y':
        pn2_has_edge = read_pickle('pn2_has_edge')
    else:
        pn2_has_edge = []
        for item in list(per_list):
            #print(item)
            if nx.has_path(G, item[0], item[1]) and item not in mix_permuation_pair:
                # if node0 -> node1 has path
                # mix_permutation_pair are impossible to construct to a path
                if nx.has_path(G, item[1], item[0]):
                    # 如果node1 -> node0(反向)有path, 而且shortest path長度為2~3之間(可自行調整數字3)
                    short_tmp = nx.shortest_path(G, item[1], item[0])
                    if len(short_tmp) >= 2 and len(short_tmp) <= 3:
                        cou = 0
                        for all_path in nx.all_simple_paths(G, item[1], item[0]):
                            # 計算所有路徑, 如果只有唯一路徑, cou == 1, 否則 cou >= 2
                            if cou >= 2:
                                break
                            cou += 1
                        if cou == 1:
                            # 如果node1 -> node0(反向)只存在唯一路徑, 代表node0 -> node1(正向)不可能找的到path通過 output -> node0 -> node1-> input
                            #print(len(short_tmp), short_tmp)
                            continue
                
                pn2_has_edge.append([item[0], item[1]])

        dump_pickle('pn2_has_edge', pn2_has_edge)
    
    print('permutation pairs num: ' + str(len(data_phase_permutation_nodes) * len(data_phase_permutation_nodes)-1) + ', has_edge pairs num: ' + str(len(pn2_has_edge)))
    print('------------------------------------------------------------')

    ########################## find path ##########################

    if input('Read data from files?(y/n) ').lower() == 'y':
        find_path = read_pickle('find_path')
        not_found_path_node = read_pickle('not_found_path_node')
        illegal_stereo_path_node = read_pickle('illegal_stereo_path_node')
    else:
        print('Start to gen path data...')
        find_path, not_found_path_node, illegal_stereo_path_node = gen_find_path(pn2_has_edge)

        dump_pickle('find_path', find_path)
        dump_pickle('not_found_path_node', not_found_path_node)
        dump_pickle('illegal_stereo_path_node', illegal_stereo_path_node)
        print('Save data completely !!')
    
    print('------------------------------------------------------------')
    print('successful path number :', len(find_path))
    print('fail path number :', len(not_found_path_node))
    # not_found_path_node = [[203,201]]        # test legal path
    print('fail path nodes :', not_found_path_node)
    print('illegal_stereo_path number :', len(illegal_stereo_path_node))
    print('illegal_stereo_path_node :', illegal_stereo_path_node)
    print('------------------------------------------------------------')
    
    #################### non Input_Node pair ####################
    '''

    for i in range(len(not_found_path_node)):
        find_path_flag = False
        print('{:>4d}->{:<4d}: '.format(not_found_path_node[i][0], not_found_path_node[i][1]))
        print('------------------------------------------------------------')
        arr_dict = dict()
        start_time = time.time()
        for output_node in output_node_id:
            a_b = []
            if nx.has_path(G, output_node, not_found_path_node[i][0]) and nx.has_path(G, not_found_path_node[i][0], not_found_path_node[i][1]):
                # for node in nx.all_simple_paths(G, output_node, not_found_path_node[i][0]):
                #     a.append(node)
                # for node in nx.all_simple_paths(G, not_found_path_node[i][0], not_found_path_node[i][1]):
                #     b.append(node)
                a = list(nx.all_simple_paths(G, output_node, not_found_path_node[i][0]))
                b = list(nx.all_simple_paths(G, not_found_path_node[i][0], not_found_path_node[i][1]))
                for x in a:
                    for y in b:
                        tmp_mix = x + y[1:]
                        if len(set(tmp_mix)) == len(tmp_mix):
                            a_b.append(tmp_mix)
            arr_dict[output_node] = a_b
        for item in arr_dict:
            print(item, len(arr_dict[item]))

        print(str(time.time() - start_time) + ' sec')

        for output_node in output_node_id:
            print('Try Output Node :', output_node)
            for input_node in input_node_id:
                if check_per_path(G, output_node, not_found_path_node[i][0], not_found_path_node[i][1], input_node):
                    if len(arr_dict[output_node]) == 0:
                        continue
                    start_time = time.time()
                    tmp_mix = []
                    c = list(nx.all_simple_paths(G, not_found_path_node[i][1], input_node))
                    # c = []
                    # for node in nx.all_simple_paths(G, not_found_path_node[i][1], input_node):
                    #     c.append(node)
                    
                    print(str(time.time() - start_time) + ' sec   ', input_node, len(c))
                    start_time = time.time()
                    cou = 0
                    for x in arr_dict[output_node]:
                        for y in c:
                            tmp_mix = x + y[1:]
                            if len(set(tmp_mix)) == len(tmp_mix):
                                print('find !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')
                                find_path_flag = True
                                break
                        cou += 1
                        # if cou%100 == 0:
                        #     print(cou)
                        if find_path_flag == True:
                            break
                    print(str(time.time() - start_time) + ' sec')
                    if find_path_flag == False:
                        continue
                    print(tmp_mix)
                    find_path_flag, tmp_mix_stereo = find_stereo_path(G, components_info, input_node_id, input_output_stereo_dict, mux_uniq_stereo_dict, mux_duplicate_stereo_dict, mux_non_pair_stereo_dict, tmp_mix)
                    if find_path_flag == False:
                        continue
                    find_path.append([tmp_mix, tmp_mix_stereo])
                    find_path_len.append(len(tmp_mix))
                    find_path_flag = True
                    break
            if find_path_flag == True:
                break
        #exit()
        if find_path_flag == False:
            print('illegal path : ', end='')
            print('{:>4d}->{:<4d}'.format(not_found_path_node[i][0], not_found_path_node[i][1]))

    '''

    #################### Input_Node pair ####################

    # ????????????



    ################### pick path(greedy) ###################

    if input('Read greedy choose path from files?(y/n) ').lower() == 'y':
        greedy_choose_path = read_pickle('greedy_choose_path')
        uncover_pairs = read_pickle('uncover_pairs')
        pn2_has_edge_covered = read_pickle('pn2_has_edge_covered')
        pattern_combination_idx = read_pickle('pattern_combination_idx')
    else:
        print('Start to gen greedy path data...')
        greedy_choose_path, uncover_pairs = greedy_pick_path(find_path, pn2_has_edge)


        pn2_has_edge_covered = []
        for i in range(len(pn2_has_edge)):
            flag = 0
            for item in uncover_pairs:
                if pn2_has_edge[i] == item:
                    flag = 1
                    break
            if flag == 0:
                pn2_has_edge_covered.append(pn2_has_edge[i])

        
        pattern_combination_idx = []
        for path_idx in range(len(greedy_choose_path)):
            pn2_idx = []
            for pair_idx in range(len(pn2_has_edge_covered)):
                mono_0_path = greedy_choose_path[path_idx][0]
                for mono_0_pre in range(len(mono_0_path)):
                    if pn2_has_edge_covered[pair_idx][0] == mono_0_path[mono_0_pre]:
                        for mono_0_post in range(mono_0_pre+1, len(mono_0_path)):
                            if pn2_has_edge_covered[pair_idx][1] == mono_0_path[mono_0_post]:
                                #print('mono_0 find pair :', pair_idx, pn2_has_edge_covered[pair_idx])
                                pn2_idx.append(pair_idx)

                mono_1_path = greedy_choose_path[path_idx][1]
                for mono_1_pre in range(len(mono_1_path)):
                    if pn2_has_edge_covered[pair_idx][0] == mono_1_path[mono_1_pre]:
                        for mono_1_post in range(mono_1_pre+1, len(mono_1_path)):
                            if pn2_has_edge_covered[pair_idx][1] == mono_1_path[mono_1_post]:
                                #print('mono_1 find pair :', pair_idx, pn2_has_edge_covered[pair_idx])
                                pn2_idx.append(pair_idx)
            
            pattern_combination_idx.append(pn2_idx)


        dump_pickle('greedy_choose_path', greedy_choose_path)
        dump_pickle('uncover_pairs', uncover_pairs)
        dump_pickle('pn2_has_edge_covered', pn2_has_edge_covered)
        dump_pickle('pattern_combination_idx', pattern_combination_idx)
        print('Save data completely !!')
    
    print('------------------------------------------------------------')
    print('pattern num :', len(greedy_choose_path))
    print('pn2_has_edge_covered :', len(pn2_has_edge_covered))
    print('--------------------- uncovered pairs ----------------------')
    print(uncover_pairs)
    print('uncovered pairs num :', len(uncover_pairs))


    ####################### dump data phase pattern log #######################
    path_log_str = ''
    for path_idx in range(len(greedy_choose_path)):
        path_log_str += 'PATH: ' + str(path_idx) + '  combination len: ' + str(len(pattern_combination_idx[path_idx])) + '\n'
        for i in range(len(pattern_combination_idx[path_idx])):
            c_idx = pattern_combination_idx[path_idx][i]
            path_log_str += 'C_' + str(pn2_has_edge_covered[c_idx][0]) + '_' + str(pn2_has_edge_covered[c_idx][1]) + ' : ' + str(c_idx) + '\n'
    with open(CHECK_LOG_DIR + '/phase_log.txt', 'w') as f:
        f.write(path_log_str)


    print('------------------------------------------------------------')
    if cust_audio_data_phase_coverage_model_changed:
        gen_coverage_data_phase(pn2_has_edge_covered)
        print('auto gen new cust_audio_data_phase_coverage_model.sv file')
    else:
        print('cust_audio_data_phase_coverage_model.sv didnt change')


    control = input("Please edit the reg_info.xlsx and in_info.xlsx out_info.xlsx then press y to gen patterns: ")
    if control.lower() == 'y':
        print('pattern_auto_gen...')
        pattern_auto_gen(greedy_choose_path, components_info, pattern_combination_idx)
        print('done')
    
    control = input('coverage self check, y or n:')
    if control.lower() == 'y':
        cov_self_check()

    ################################################################## main ##################################################################