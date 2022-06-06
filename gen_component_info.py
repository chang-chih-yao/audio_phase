import os
import json
import re
import pandas as pd
from colorama import Fore, Style
from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import math

########################## global settings ##########################
SIGNATURE = '/*********** Howard Auto Gen Tools ***********/\n'
COMPONENT_NAME_COLUMN_NAME = '主圖形名稱'
BACKUP_DIR  = 'backup'
INPUT_DIR   = 'input'
ENV_DIR     = 'env'
CONTENT_DIR = 'env_content'
DATA_DIR    = 'data'
PATTERN_DIR = 'audio_data_path_auto_gen_patterns'
CHECK_LOG_DIR = 'check_log'
REPORT_DIR  = 'report'
NULL_NODE_ID = -1
TAB = '    '
SELECT_SIGNAL_RESERVED = 'reserved'
REG_READ_WRITE_CMD_ADDR = 'UVM_ADDR'
REG_READ_WRITE_CMD_DATA = 'UVM_DATA'
REG_READ_WRITE_VARIABLE = 'UVM_VARIABLE'
REG_READ_WRITE_VARIABLE_NAME = 'data_tmp'
RESERVED_SELECT_SIGNAL = SELECT_SIGNAL_RESERVED
OFFSET_SELECT_SIGNAL   = '_offset_'
FOR_SD_CHECK_ONLY = False

########################## global settings ##########################

def WARNING_LOG(LOG):
    return '[WARNING] ' + LOG

def ERROR_LOG(LOG):
    return '[ERROR] ' + LOG

def COLOR_MESSAGE(message, SETTINGS):
    #init()
    #print(SETTINGS + message + Style.RESET_ALL)
    print(message)

def PASS_MESSAGE(message = 'PASS\n'):
    SETTINGS = Style.BRIGHT + Fore.GREEN
    COLOR_MESSAGE(message, SETTINGS)

def FAIL_MESSAGE(message = 'FAIL\n'):
    SETTINGS = Style.BRIGHT + Fore.RED
    COLOR_MESSAGE(message, SETTINGS)

def ERROR_MESSAGE(message):
    SETTINGS = Style.BRIGHT + Fore.RED
    COLOR_MESSAGE(ERROR_LOG(message), SETTINGS)

def WARNING_MESSAGE(message):
    SETTINGS = Style.BRIGHT + Fore.YELLOW
    COLOR_MESSAGE(WARNING_LOG(message), SETTINGS)


def contents_append_endl(CONTENTS, tmp):
    CONTENTS.append(tmp + '\n')

def contents_append_tab_endl(CONTENTS, tmp):
    CONTENTS.append(TAB + tmp + '\n')

def contents_append_tab_tab_endl(CONTENTS, tmp):
    CONTENTS.append(TAB + TAB + tmp + '\n')

def contents_append_tab_tab_tab_endl(CONTENTS, tmp):
    CONTENTS.append(TAB + TAB + TAB + tmp + '\n')

def file_with_dir_path(output_file_name, output_dir):
    os.makedirs(output_dir, exist_ok=True)
    return os.path.join(output_dir, output_file_name)

def write_contents_to_file(CONTENTS, output_file_name, output_dir, IS_APPEND_MODE = False, ENDL = '\n'):
    output_file_name = file_with_dir_path(output_file_name, output_dir)
   
    if IS_APPEND_MODE:
        output_file = open(output_file_name, "a", encoding = "UTF-8")
    else:
        output_file = open(output_file_name, "w", encoding = "UTF-8")

    for content in CONTENTS:
        output_file.write(content + ENDL)
    output_file.close()

def get_components_parsing_rule(input_file_name = 'components_parsing_rule.json', input_dir = INPUT_DIR):
    input_file_name = file_with_dir_path(input_file_name, input_dir)
    
    with open(input_file_name) as json_file: 
        components_parsing_rule = json.load(json_file)
        
    return components_parsing_rule

def fill_component_info(one_component_dict, component, i, component_data_frame, IN_PATTERN, OUT_PATTERN, SELECT_PATTERN, SPLIT_POSTFIX = ''):    
    ## parse Out
    try:
        columns_list = list(filter(re.compile(OUT_PATTERN).match, component_data_frame.columns)) 
        output_list = []
        for column in columns_list:
            if str(component_data_frame.iloc[i][column]) != 'nan':
                output = component_data_frame.iloc[i][column].lower()
                output_list.append(output)
                #print('{}: {}'.format(column, component_data_frame.iloc[i][column]))
        one_component_dict.update({'Outputs': output_list})
    except:
        #traceback.print_exc()
        #print('No Out')
        one_component_dict.update({'Outputs': []})
        
    ## case 0: there is a reserved select signal
    ## case 1: there is n-1 offset select signal and 1 non-offset select signal
    CASE0 = False
    CASE1 = False
    CASE0_INPUT_IDX = -1
    CASE1_INPUT_IDX = -1
    num_offset_select_signal     = 0
    num_non_offset_select_signal = 0
    
    ## parse Select
    try:
        columns_list = list(filter(re.compile(SELECT_PATTERN).match, component_data_frame.columns))
        select_list = []
        
        ## check case 0
        for idx, column in enumerate(columns_list):
            if str(component_data_frame.iloc[i][column]) != 'nan':
                select = component_data_frame.iloc[i][column].lower()
                if select == RESERVED_SELECT_SIGNAL:
                    CASE0_INPUT_IDX = idx
                    CASE0 = True
                else:
                    if select.find(OFFSET_SELECT_SIGNAL) != -1:
                        num_offset_select_signal += 1
                    else:
                        num_non_offset_select_signal += 1
                        CASE1_INPUT_IDX = idx
        
        for idx, column in enumerate(columns_list):
            if str(component_data_frame.iloc[i][column]) != 'nan':
                select = component_data_frame.iloc[i][column].lower()
                select_list.append(select)

        #print(num_offset_select_signal)
        #print(num_non_offset_select_signal)
        #print('\n')
        if num_non_offset_select_signal == 1 and num_offset_select_signal > 0 and num_offset_select_signal + num_non_offset_select_signal == len(select_list):
            CASE1 = True
            
        #if CASE0:
        #    print('CASE0')
        #    print(select_list)
        
        #if CASE1:
        #    print('CASE1')
        #    print(select_list)

        if CASE0:
            one_component_dict.update({'Not_Covered_Select': select_list})
            select_list = []
            one_component_dict.update({'Type': one_component_dict['Type'] + '_with_reserved'})
        elif CASE1:
            one_component_dict.update({'Not_Covered_Select': select_list})
            select_list = []
            one_component_dict.update({'Type': one_component_dict['Type'] + '_with_offset'})
            
        one_component_dict.update({'Selects': select_list})
        
        if str(component['Select']).find('Mute') != -1:
            one_component_dict.update({'IS_REVERSE': True})
        else:
            one_component_dict.update({'IS_REVERSE': False})
            
    except:
        #traceback.print_exc()
        #print('No Select')
        one_component_dict.update({'Selects': []})
        
            
    ## parse In
    try:
        columns_list = list(filter(re.compile(IN_PATTERN).match, component_data_frame.columns))
        input_list = []
        not_covered_input_list = []
        for idx, column in enumerate(columns_list):
            if str(component_data_frame.iloc[i][column]) != 'nan':
                input = component_data_frame.iloc[i][column].lower()
                if CASE0:
                    if idx == CASE0_INPUT_IDX:
                        input_list.append(input)
                    else:
                        not_covered_input_list.append(input)
                elif CASE1:
                    if idx == CASE1_INPUT_IDX:
                        input_list.append(input)
                    else:
                        not_covered_input_list.append(input)
                else:
                    input_list.append(input)
                #print('{}: {}'.format(column, component_data_frame.iloc[i][column]))
        one_component_dict.update({'Inputs': input_list})
        if CASE0 or CASE1:
            one_component_dict.update({'Not_Covered_Inputs': not_covered_input_list})
    except:
        #traceback.print_exc()
        #print('No In')
        one_component_dict.update({'Inputs': []})

    return one_component_dict

def regex_filter(val, regex):
    if val:
        try:
            mo = re.search(regex, val)
        except:
            mo = False
        if mo:
            return True
        else:
            return False
        
    else:
        return False

def get_components_info(components_parsing_rule, input_file_name = 'all.xlsx', input_dir = INPUT_DIR, replace_dot = True):
    print('--------- Parsing Start ---------')
    components_file = file_with_dir_path(input_file_name, input_dir)
    excel_df = pd.read_excel(components_file, skiprows=[0])
    df = pd.DataFrame(excel_df)
    components_info = []
    
    total_got = 0

    for component in components_parsing_rule:
        print("components_info length: {}".format(len(components_info)))
        component_data_frame = df.loc[df[COMPONENT_NAME_COLUMN_NAME].apply(regex_filter, regex=component['Name'])]
        total_got += len(component_data_frame)
        print('Got {} {}'.format(len(component_data_frame), component['Name']))
        for i in range(0, len(component_data_frame)):
            type = component_data_frame.iloc[i][COMPONENT_NAME_COLUMN_NAME]
            if replace_dot:
                type = re.sub(r"\.[0-9]+", "", type)
            #print(component_data_frame.iloc[i][COMPONENT_NAME_COLUMN_NAME])
            
            if 'MulOut_NoSel' in component:
                for idx, split in enumerate(component['MulOut_NoSel']):
                    one_component_dict = {'Type': type + '_' + split, 'IS_REVERSE': False}
                    one_component_dict = fill_component_info(one_component_dict, component, i, component_data_frame, component['In'][idx], component['Out'][idx], component['Select'], SPLIT_POSTFIX = split)
                    components_info.append(one_component_dict)
            else:    
                one_component_dict = {'Type': type, 'IS_REVERSE': False}
                one_component_dict = fill_component_info(one_component_dict, component, i, component_data_frame, component['In'], component['Out'], component['Select'])
                components_info.append(one_component_dict)

    print('--------- Parsing Summary: got total {} components ---------\n'.format(total_got))
    return components_info

def make_distinct(components_info):
    new_components_info = components_info.copy()
    #for idx, x in enumerate(new_components_info):
    #    for y in new_components_info[idx+1:]:
    #        if x['Type'] == y['Type'] and x['Inputs'] == y['Inputs'] and x['Outputs'] == y['Outputs'] and x['Selects'] == y['Selects']:
    #            new_components_info.remove(y)
    #            
    #WARNING_MESSAGE("make dinstinct result, ori({}), new({})".format(len(components_info), len(new_components_info)))
    return new_components_info

def to_lower(components_info):
    for component_info in components_info:
        for input in component_info['Inputs']:
            input = input.lower()
            
        for output in component_info['Outputs']:
            output = output.lower()
              
        for select in component_info['Selects']:
            select = select.lower()
    return components_info

def sort_components_info(components_info):
    newlist = sorted(components_info, key=lambda k: k['Type']) 
    return newlist

def set_components_id(components_info):
    for idx, component_info in enumerate(components_info):
        component_info.update({"NODE_ID": str(idx)})
        
    return components_info

def output_components_info(components_info, output_file_name = 'components_info.txt', output_dir = DATA_DIR):
    CONTENTS = []
    for idx, component_info in enumerate(components_info):
        contents_append_endl(CONTENTS, 'Component[%3s]: ' % str(idx) + json.dumps(component_info))
        
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def parse_register_file(input_file_name = 'all_register_dump_for_dv.txt', input_dir = INPUT_DIR):
    all_register_info = {}
    input_file_name = file_with_dir_path(input_file_name, input_dir)
    
    try:
        all_register_dump_for_dv_file = open(input_file_name, 'r')
    except:
        ERROR_MESSAGE('all_register_dump_for_dv.txt not found')
        control = input("Please prepare all_register_dump_for_dv.txt in the input directory, and press C(continue)")
        if control.lower() != 'c':
            return all_register_info
    
    for idx, line in enumerate(all_register_dump_for_dv_file.readlines()):
        if idx != 0:
            split_result = re.split('\s+', line)
            Signal = split_result[0]
            Addr   = split_result[1]
            Bits_Indices = split_result[2]
            
            delta = -1
            if Bits_Indices.find(':') != -1:
                start = int(Bits_Indices[:Bits_Indices.find(':')])
                end   = int(Bits_Indices[Bits_Indices.find(':')+1:])
                delta = int(abs(end-start))+1
            else:
                delta = 1
            
            all_register_info.update({ Signal: {'Addr': Addr, 'Bits_Indices': Bits_Indices, 'Length': delta}})
    
    return all_register_info
    
def create_excel_table_title(ws, col_names, color = "EBBA34", with_border = True):
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    ## the col names
    ws.append(col_names)
    for cell in ws[1]:
        cell.fill = PatternFill(start_color="EBBA34", fill_type = "solid")
        #cell.font = Font(bold=True)
        cell.border = thin_border
        
    return ws

def excel_auto_width(ws):
    EXTENSION = 2
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + EXTENSION
    return ws

def construct_registed_info_sheet(all_register_info, wb, sheet_name=''):
    if not sheet_name:
        ws = wb.active
    else:
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

    col_names = ['Signal', 'Addr', 'Bits_Indices', 'Length']
    rows = []

    for Signal in all_register_info:
        register = all_register_info[Signal]
        row = []
        row.append(Signal)
        for col_name in col_names[1:]:
            row.append(register[col_name])

        rows.append(row)
        
    
    ws = create_excel_table_title(ws, col_names)

    for row in rows:
        ws.append(row)
        
    ws = excel_auto_width(ws)

def output_register_info(all_register_info, output_file_name = 'all_register_info.xlsx', output_dir = DATA_DIR):
    wb = Workbook()
    
    construct_registed_info_sheet(all_register_info, wb)
    
    output_file_name = file_with_dir_path(output_file_name, output_dir)
    wb.save(output_file_name)

def check_single_output(components_info, output_file_name = 'check_single_output.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    for idx, component_info in enumerate(components_info):
        if len(component_info['Outputs']) > 1:
            PASS = False
            component_id   = '%3s'  % idx
            component_type = '%35s' % component_info['Type']
            
            outputs = ''
            for output in component_info['Outputs']:
                outputs += output + ' '

            LOG = 'Error: : component[{}]({}) has multiple outputs: {}'.format(component_id, component_type, outputs)
            CONTENTS.append(LOG)
            print(LOG)
            
    write_contents_to_file(CONTENTS, output_file_name, output_dir)        
    return PASS

def check_merge(CHECK_ALL_PASS, CHECK_PASS):
    if CHECK_PASS:
        PASS_MESSAGE()
    else:
        FAIL_MESSAGE()
    return CHECK_PASS and CHECK_ALL_PASS

def check_output_distinct(components_info, output_file_name = 'check_output_distinct.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    for idx_x, x in enumerate(components_info):
        for idx_y, y in enumerate(components_info):
            if idx_x < idx_y:
                for x_output in x['Outputs']:
                    for y_output in y['Outputs']:
                        if x_output == y_output:
                            PASS = False
                            component_id_x   = '%3s'  % idx_x
                            component_type_x = '%35s' % x['Type']
                            component_id_y   = '%3s'  % idx_y
                            component_type_y = '%35s' % y['Type']
                            output_signal    = '%35s' % x_output
                            LOG = 'Error: component[{}]({}) and component[{}]({}) output: {} are the same'.format(component_id_x, component_type_x, component_id_y, component_type_y, output_signal)
                            ERROR_MESSAGE(LOG)
                            CONTENTS.append(ERROR_LOG(LOG))
                            
    write_contents_to_file(CONTENTS, output_file_name, output_dir)      
    return PASS

def check_output_connection(components_info, output_file_name = 'check_output_connection.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    for idx, component_info in enumerate(components_info):
        if len(component_info['Outputs']) == 1 :
            output = component_info['Outputs'][0]
            Found = False
            for idx_y, y in enumerate(components_info):
                for input in y['Inputs']:
                    if input == output:
                        Found = True
                        break
            if not Found:
                PASS = False
                component_id   = '%3s'  % idx
                component_type = '%35s' % component_info['Type']
                output_signal  = '%35s' % output
                LOG = 'no_connection: output({}) of component[{}]({})'.format(output_signal, component_id, component_type)
                CONTENTS.append(ERROR_LOG(LOG))
                ERROR_MESSAGE(LOG)
        
    write_contents_to_file(CONTENTS, output_file_name, output_dir)
    return PASS

def check_input_connection(components_info, output_file_name = 'check_input_connection.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    input_signals = [SELECT_SIGNAL_RESERVED]
    
    for input_node in components_info:
        if input_node['Type'] == 'Input_Node':
            input_signals.append(input_node['Outputs'][0])
            
    for idx_y, y in enumerate(components_info):
        for input in y['Inputs']:
            if input not in input_signals:
                Found = False
                for idx_x, x in enumerate(components_info):
                    if len(x['Outputs']) == 1 :
                        output = x['Outputs'][0]
                        if input == output:
                            Found = True
                            break
                            
                if not Found:
                    PASS = False
                    component_id   = '%3s'  % idx_y
                    component_type = '%35s' % y['Type']
                    input_signal   = '%45s'  % input
                    LOG = 'no_connection: input({}) of component[{}]({})'.format(input_signal, component_id, component_type)
                    WARNING_MESSAGE(LOG)
                    CONTENTS.append(WARNING_LOG(LOG))
        
    write_contents_to_file(CONTENTS, output_file_name, output_dir)
    #return PASS
    return True

def check_select_and_input_num(components_info, output_file_name = 'check_select_and_input_num.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    
    for idx, component_info in enumerate(components_info):
        input_num  = len(component_info['Inputs'])
        select_num = len(component_info['Selects'])
        ## 要嘛數量相等 要嘛只有一個 mul-bit select siganl
        if (input_num != select_num) and not (input_num == 1 and select_num == 0):
            if select_num != 1:
                PASS = False
                component_id   = '%3s'  % idx
                component_type = '%35s' % component_info['Type']
                component_input_num  = '%3s' % input_num
                component_select_num = '%3s' % select_num
                LOG = 'INPUT NUM and SELECT NUM CONFLICT: component[{}]({}) input_num = {}, select_num = {}'.format(component_id, component_type, component_input_num, component_select_num)
                CONTENTS.append(LOG)
                print(LOG)
                
    write_contents_to_file(CONTENTS, output_file_name, output_dir)
    return PASS

def check_select_in_all_register_info(components_info, all_register_info, output_file_name = 'check_select_in_all_register_info.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    
    for idx, component_info in enumerate(components_info):
        for select_signal in component_info['Selects']:
            if select_signal != SELECT_SIGNAL_RESERVED and select_signal not in all_register_info:
                component_select_signal = '%35s' % select_signal
                PASS = False
                LOG = 'select signal({}) not found in all_register_info'.format(component_select_signal)
                CONTENTS.append(ERROR_LOG(LOG))
                ERROR_MESSAGE(LOG)  

    write_contents_to_file(CONTENTS, output_file_name, output_dir)
    return PASS

def check_select_distinct(components_info, output_file_name = 'check_select_distinct.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    for idx_x, x in enumerate(components_info):
        for idx_y, y in enumerate(components_info):
            if idx_x < idx_y:
                for x_select in x['Selects']:
                    for y_select in y['Selects']:
                        if x_select == y_select:
                            PASS = False
                            component_id_x   = '%3s'  % idx_x
                            component_type_x = '%35s' % x['Type']
                            component_id_y   = '%3s'  % idx_y
                            component_type_y = '%35s' % y['Type']
                            select_signal    = '%35s' % x_select
                            LOG = 'component[{}]({}) and component[{}]({}) select: {} are the same'.format(component_id_x, component_type_x, component_id_y, component_type_y, select_signal)
                            CONTENTS.append(WARNING_LOG(LOG))
                            #WARNING_MESSAGE(LOG)
                            
    write_contents_to_file(CONTENTS, output_file_name, output_dir)      
    return PASS

def check_pow_of_2(n):
    return n & (n-1) == 0

def check_components_info_and_all_register_info(components_info, all_register_info, output_file_name = 'check_components_info_and_all_register_info.txt', output_dir = CHECK_LOG_DIR):
    PASS = True
    CONTENTS = []
    
    for idx, component_info in enumerate(components_info):
        input_num  = len(component_info['Inputs'])
        select_num = len(component_info['Selects'])
        
        ## 確認 mul-bit select siganl
        if input_num != select_num and not (input_num == 1 and select_num == 0) and select_num == 1:
            select_signal = component_info['Selects'][0]
            component_select_signal = '%35s' % select_signal
            if select_signal in all_register_info:
                signal_width  = all_register_info[select_signal]['Length']
                component_id   = '%3s'  % idx
                component_type = '%35s' % component_info['Type']
                component_input_num    = '%3s' % input_num
                component_signal_width = '%3s' % signal_width

                LOG = 'input num and select signal({}) width conflict: component[{}]({}) input_num = {}, signal_width = {}'.format(component_select_signal, component_id, component_type, component_input_num, component_signal_width)
                
                if input_num > math.pow(2, signal_width):
                    PASS = False
                    CONTENTS.append(ERROR_LOG(LOG))
                    ERROR_MESSAGE(LOG)
                elif input_num != math.pow(2, signal_width) and check_pow_of_2(input_num):
                    CONTENTS.append(WARNING_LOG(LOG))
                    WARNING_MESSAGE(LOG)
            else:
                PASS = False
                LOG = 'select signal({}) not found in all_register_info'.format(component_select_signal)
                CONTENTS.append(ERROR_LOG(LOG))
                ERROR_MESSAGE(LOG)

    write_contents_to_file(CONTENTS, output_file_name, output_dir)
    return PASS

def output_mux_gentop(components_info, output_file_name = 'mux_gentop.v', output_dir = DATA_DIR):
    MUX_NAME_RE = r'^MUX_[1-9]*[0-9]to1'
    CONTENTS = []
    for idx, component_info in enumerate(components_info):
        if re.match(MUX_NAME_RE, component_info['Type']):
            lower_name = 'dv_' + component_info['Type'].lower()
            contents_append_endl(CONTENTS, '// {}'.format(lower_name))
            
            for idx, Input in enumerate(component_info['Inputs']):
                input_format = Input.ljust(60) if Input != SELECT_SIGNAL_RESERVED else "32'h0".ljust(60)
                contents_append_endl(CONTENTS, '// realtek conn ' + input_format + '{}.i{}'.format(lower_name, idx))
                
            contents_append_endl(CONTENTS, '// realtek conn ' + component_info['Selects'][0].ljust(60) + '{}.sel'.format(lower_name))
            contents_append_endl(CONTENTS, '// realtek conn ' + component_info['Outputs'][0].ljust(60) + '{}.out'.format(lower_name))
            
        
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def make_legal(components_info, ilegal_pattern = r"[^a-z0-9_]", replace_with = '_'):
    for component_idx, component_info in enumerate(components_info):
        for idx, input in enumerate(component_info['Inputs']):
            old = input
            new = re.sub(ilegal_pattern, replace_with, old, count=100)
            if old != new:
                WARNING  = 'component[%3s] input  naming illegal, ' % component_idx
                WARNING += 'auto changed from %35s' % old
                WARNING += ' to %35s' % new
                WARNING_MESSAGE(WARNING)
                component_info['Inputs'][idx] = new
            
        for idx, output in enumerate(component_info['Outputs']):
            old = output
            new = re.sub(ilegal_pattern, replace_with, old, count=100)
            if old != new:
                WARNING  = 'component[%3s] output naming illegal, ' % component_idx
                WARNING += 'auto changed from %35s' % old
                WARNING += ' to %35s' % new
                WARNING_MESSAGE(WARNING)
                component_info['Outputs'][idx] = new
              
        for idx, select in enumerate(component_info['Selects']):
            old = select
            new = re.sub(ilegal_pattern, replace_with, old, count=100)
            if old != new:
                WARNING  = 'component[%3s] select naming illegal, ' % component_idx
                WARNING += 'auto changed from %35s' % old
                WARNING += ' to %35s' % new
                WARNING_MESSAGE(WARNING)
                component_info['Selects'][idx] = new
            
    return components_info

def get_nodes_by_type(components_info, type):
    nodes = []
    for component_info in components_info:
        if component_info['Type'] == type:
            nodes.append(component_info)
            
    return nodes

def gen_content_of_cust_system_configuration(components_info, output_file_name = 'cust_system_configuration.sv', output_dir = CONTENT_DIR):
    CONTENTS = []
    contents_append_endl(CONTENTS, SIGNATURE)
    input_nodes  = get_nodes_by_type(components_info, 'Input_Node')
    output_nodes = get_nodes_by_type(components_info, 'Output_Node')
    
    contents_append_tab_endl(CONTENTS, '/************ audio data path settings begin ************/')
    contents_append_tab_endl(CONTENTS, 'this.audio_data_path_cfg[0].source_num ' + '%19s' % '= {};'.format(len(input_nodes)) + '// # of input node')
    contents_append_tab_endl(CONTENTS, 'this.audio_data_path_cfg[0].channel_num' + '%19s' % '= {};'.format(len(output_nodes)) + '// # of output node')
    contents_append_tab_endl(CONTENTS, 'this.audio_data_path_cfg[0].create_sub_cfgs();')
    contents_append_endl(CONTENTS, '')
    
    contents_append_tab_endl(CONTENTS, '// input nodes')
    for idx, input_node in enumerate(input_nodes):
        contents_append_tab_endl(CONTENTS, 'this.audio_data_path_cfg[0].source_name[%3s]' % idx + '%11s' % '= ' + '"{}";'.format(input_node['Outputs'][0]))
    contents_append_endl(CONTENTS, '')
    
    contents_append_tab_endl(CONTENTS, '// output nodes')
    for idx, output_node in enumerate(output_nodes):
        contents_append_tab_endl(CONTENTS, 'this.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].channel_name' % idx + ' = "{}";'.format(output_node['Inputs'][0]))
    contents_append_endl(CONTENTS, '')
    
    contents_append_tab_endl(CONTENTS, '/************* audio data path settings end *************/')
    
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_audio_data_path_golden_pattern(components_info, output_file_name = 'audio_data_path_golden.sv', output_dir = CONTENT_DIR):
    CONTENTS = []
    contents_append_endl(CONTENTS, SIGNATURE)
    output_nodes = get_nodes_by_type(components_info, 'Output_Node')
    
    contents_append_tab_endl(CONTENTS, '/************ audio data path settings begin ************/')
    contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_enable = 1;')
    contents_append_endl(CONTENTS, '')
    
    for idx, output_node in enumerate(output_nodes):
        contents_append_tab_endl(CONTENTS, '// performance check setting of output: {}'.format(output_node['Inputs'][0]))
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].channel_enable     = 0;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].is_dwa             = 0;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].sample_bit         = AUDIO_DATA_PATH_SAMPLE_BIT_16;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].sample_rate        = AUDIO_DATA_PATH_SAMPLE_RATE_48K;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].fs                 = 6144000;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].sample_num         = 320000;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].discard_sample_num = 128000;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].thd_threshold      = -90;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].amp_threshold      = -8.5;' % idx)
        contents_append_endl(CONTENTS, '')
    
    contents_append_tab_endl(CONTENTS, '/************* audio data path settings end *************/')
    
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_audio_data_phase_golden_pattern(components_info, output_file_name = 'audio_data_phase_golden.sv', output_dir = CONTENT_DIR):
    CONTENTS = []
    contents_append_endl(CONTENTS, SIGNATURE)
    output_nodes = get_nodes_by_type(components_info, 'Output_Node')
    
    contents_append_tab_endl(CONTENTS, '/************ audio data path settings begin ************/')
    contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_enable = 1;')
    contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_phase_enable = 1;')
    contents_append_endl(CONTENTS, '')
    
    for idx, output_node in enumerate(output_nodes):
        contents_append_tab_endl(CONTENTS, '// performance check setting of output: {}'.format(output_node['Inputs'][0]))
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].channel_enable     = 0;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].is_dwa             = 0;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].sample_bit         = AUDIO_DATA_PATH_SAMPLE_BIT_16;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].sample_rate        = AUDIO_DATA_PATH_SAMPLE_RATE_48K;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].fs                 = 6144000;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].sample_num         = 320000;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].discard_sample_num = 128000;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].thd_threshold      = -90;' % idx)
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].audio_data_path_channel_cfg[%3s].amp_threshold      = -8.5;' % idx)
        contents_append_endl(CONTENTS, '')
    
    contents_append_tab_endl(CONTENTS, '/************* audio data path settings end *************/')
    
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_transition_model(components_info, output_file_name = 'cust_audio_data_path_transition_model.sv', output_dir = ENV_DIR, output_enable = True):
    CONTENTS = []
    contents_append_endl(CONTENTS, SIGNATURE)
    ###################################### class init ######################################
    contents_append_endl(CONTENTS, '`ifndef CUST_AUDIO_DATA_PATH_TRANSITION_MODEL__SV')
    contents_append_endl(CONTENTS, '`define CUST_AUDIO_DATA_PATH_TRANSITION_MODEL__SV')
    contents_append_endl(CONTENTS, 'class cust_audio_data_path_transition_model extends audio_data_path_transition_model #(cust_audio_data_path_dump_mux_transaction);')
    contents_append_tab_endl(CONTENTS, '`uvm_object_utils_begin(cust_audio_data_path_transition_model)')
    contents_append_tab_endl(CONTENTS, '`uvm_object_utils_end')
    contents_append_tab_endl(CONTENTS, 'extern function new(string name = "cust_audio_data_path_transition_model");')
    contents_append_tab_endl(CONTENTS, 'extern function int transition(ref cust_audio_data_path_dump_mux_transaction tr, int destination_id);')
    contents_append_tab_endl(CONTENTS, 'extern function int selection(ref cust_audio_data_path_dump_mux_transaction tr, int node_id);')
    contents_append_endl(CONTENTS, 'endclass: cust_audio_data_path_transition_model')
    ###################################### class init ######################################
    
    contents_append_endl(CONTENTS, "")
    
    ####################################### tree init #######################################
    contents_append_endl(CONTENTS, 'function cust_audio_data_path_transition_model::new(string name = "cust_audio_data_path_transition_model");')
    contents_append_tab_endl(CONTENTS, 'super.new(name);')
    
    contents_append_tab_endl(CONTENTS, 'select_tree = new[{}](select_tree);'.format(len(components_info)))
    for component_info in components_info:
        INPUT_NUM        = len(component_info['Inputs'])
        SELECT_TREE_NODE = 'select_tree[{}]'.format(component_info['NODE_ID'])
        
        contents_append_endl(CONTENTS, '\n' + TAB + '// Type:   {}'.format(component_info['Type']))
        for input in component_info['Inputs']:
            contents_append_tab_endl(CONTENTS, '// Input:  {}'.format(input))
        for output in component_info['Outputs']:
            contents_append_tab_endl(CONTENTS, '// Output: {}'.format(output))
        
        if INPUT_NUM > 0:
            contents_append_tab_endl(CONTENTS, '{}.select = new[{}];'.format(SELECT_TREE_NODE, INPUT_NUM))
            contents_append_tab_endl(CONTENTS, '{}.select_done = new[{}];'.format(SELECT_TREE_NODE, INPUT_NUM))
            for i in range(INPUT_NUM):
                contents_append_tab_endl(CONTENTS, '{}.select[{}] = -1;'.format(SELECT_TREE_NODE, str(i)))
            for i in range(INPUT_NUM):
                contents_append_tab_endl(CONTENTS, "{}.select_done[{}] = 1'b0;".format(SELECT_TREE_NODE, str(i)))
        else:
            contents_append_tab_endl(CONTENTS, '{}.select = new[{}];'.format(SELECT_TREE_NODE, 1))
            contents_append_tab_endl(CONTENTS, '{}.select_done = new[{}];'.format(SELECT_TREE_NODE, 1))
            contents_append_tab_endl(CONTENTS, '{}.select[{}] = -2;'.format(SELECT_TREE_NODE, 0))
            contents_append_tab_endl(CONTENTS, "{}.select_done[{}] = 1'b0;".format(SELECT_TREE_NODE, 0))
            
        contents_append_tab_endl(CONTENTS, '{}.source = -1;'.format(SELECT_TREE_NODE))
    ####################################### tree init #######################################

    contents_append_endl(CONTENTS, "\n")    
        
    #################################### tree connection ####################################
    for child in components_info:
        SELECT_TREE_NODE = 'select_tree[{}]'.format(child['NODE_ID'])

        child_select_parents_node_id = []
        child_input_cnt = 0
        for input in child['Inputs']:
            INPUT_FOUND = False
            for parent in components_info:
                if len(parent['Outputs']) > 0:
                    if input == parent['Outputs'][0]:
                        child_select_parents_node_id.append(int(parent['NODE_ID']))
                        contents_append_tab_endl(CONTENTS, '{}.select[{}] = {}; //connection: {}'.format(SELECT_TREE_NODE, str(child_input_cnt), parent['NODE_ID'], input))
                        INPUT_FOUND = True
            if not INPUT_FOUND:
                child_select_parents_node_id.append(NULL_NODE_ID)
                
            child_input_cnt += 1
        child.update({'select': child_select_parents_node_id})
    #################################### tree connection ####################################
    
    ################################### source connection ###################################
    NUM_INPUT_NODE = 0
    for component in components_info:
        if component['Type'] == 'Input_Node':
            SELECT_TREE_NODE = 'select_tree[{}]'.format(component['NODE_ID'])
            output = component['Outputs'][0]
            contents_append_tab_endl(CONTENTS, '{}.source = {}; //input: {}'.format(SELECT_TREE_NODE, NUM_INPUT_NODE, output))
            NUM_INPUT_NODE += 1
    ################################### source connection ###################################
    
    contents_append_endl(CONTENTS, "\nendfunction: new\n")
    
    ###################################### tree transition ######################################
    contents_append_endl(CONTENTS, "/** transition for cust_audio_data_path_transition_model */")
    contents_append_endl(CONTENTS, 'function int cust_audio_data_path_transition_model::transition(ref cust_audio_data_path_dump_mux_transaction tr, int destination_id);')
    contents_append_tab_endl(CONTENTS, 'int select, select_tmp;')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, '`uvm_info("selection", $sformatf("destination id: %0d", destination_id), UVM_LOW)')
    contents_append_tab_endl(CONTENTS, '`uvm_info("AUDIO_DATA_PATH_DUMP_MUX", tr.sprint(), UVM_DEBUG)')
    contents_append_endl(CONTENTS, '')
    destination_id = 0
    for component in components_info:
        if component['Type'] == 'Output_Node':
            NODE_ID = component['NODE_ID']
            input = component['Inputs'][0]
            if destination_id == 0:
                contents_append_tab_endl(CONTENTS, 'if(destination_id == {})'.format(destination_id))
            else:
                contents_append_tab_endl(CONTENTS, 'else if(destination_id == {})'.format(destination_id))
            contents_append_tab_tab_endl(CONTENTS, 'select = {};//output: {}'.format(NODE_ID, input))

            destination_id += 1
    contents_append_tab_endl(CONTENTS, 'else')
    contents_append_tab_tab_endl(CONTENTS, 'select = -1;')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, 'foreach(select_tree[i]) begin')
    contents_append_tab_tab_endl(CONTENTS, 'select_tree[i].done = 0;')
    contents_append_tab_endl(CONTENTS, 'end')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, 'path_log = $sformatf("%0d", select);')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, 'while((select != -1) && (select != -2)) begin')
    contents_append_tab_tab_endl(CONTENTS, 'select_tmp = select;')
    contents_append_tab_tab_endl(CONTENTS, 'select = selection(tr, select);')
    contents_append_tab_tab_endl(CONTENTS, 'if((select != -1) && (select != -2))')
    contents_append_tab_tab_tab_endl(CONTENTS, 'path_log = {path_log, $sformatf(" %0d", select)};')
    contents_append_tab_tab_endl(CONTENTS, '`uvm_info("selection", $sformatf("select node: %0d", select), UVM_LOW)')
    contents_append_tab_endl(CONTENTS, 'end')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, 'if(select == -2) begin')
    contents_append_tab_tab_endl(CONTENTS, '`uvm_info("selection", $sformatf("select source node: %0d (%s)", select_tree[select_tmp].source, cfg.source_name[select_tree[select_tmp].source]), UVM_LOW)')
    contents_append_tab_tab_endl(CONTENTS, 'compare_path_log();')
    contents_append_tab_tab_endl(CONTENTS, 'return select_tree[select_tmp].source;')
    contents_append_tab_endl(CONTENTS, 'end')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, 'return -1;')
    contents_append_endl(CONTENTS, 'endfunction: transition')
    ###################################### tree transition ######################################
    
    contents_append_endl(CONTENTS, '')
    
    ###################################### tree selection #######################################
    contents_append_endl(CONTENTS, "/** selection for cust_audio_data_path_transition_model */")
    contents_append_endl(CONTENTS, 'function int cust_audio_data_path_transition_model::selection(ref cust_audio_data_path_dump_mux_transaction tr, int node_id);')
    contents_append_tab_endl(CONTENTS, 'if(select_tree[node_id].done == 1)')
    contents_append_tab_tab_endl(CONTENTS, 'return -1;')
    contents_append_tab_endl(CONTENTS, 'select_tree[node_id].done = 1;')
    
    for component_info in components_info:
        node_id = component_info['NODE_ID']
        contents_append_endl(CONTENTS, "")
        contents_append_tab_endl(CONTENTS, "if(node_id == {}) begin".format(node_id))
        
        select_num = len(component_info['Selects'])
        input_num  = len(component_info['Inputs'])
        output_num = len(component_info['Outputs'])
        
        if select_num == 1:
            select    = component_info['Selects'][0]
            input_num = len(component_info['Inputs'])
            for i in range(input_num):
                if component_info['IS_REVERSE']:
                    contents_append_tab_tab_endl(CONTENTS, "if( !tr.{} == {} ) begin".format(select, i))
                else:
                    contents_append_tab_tab_endl(CONTENTS, "if( tr.{} == {} ) begin".format(select, i))
                    
                input_signal = component_info['Inputs'][i]
                contents_append_tab_tab_endl(CONTENTS, TAB + '`uvm_info("selection", $sformatf("{} is {}, node[%0d] -> node[%0d] through {}", node_id, select_tree[{}].select[{}]), UVM_LOW)'.format(select, i,  input_signal, node_id, i))
                    
                contents_append_tab_tab_endl(CONTENTS, TAB + "select_tree[{}].select_done[{}] = 1'b1;".format(node_id, i))
                contents_append_tab_tab_endl(CONTENTS, TAB + "return select_tree[{}].select[{}];".format(node_id, i))
                contents_append_tab_tab_endl(CONTENTS, "end")
                
        elif select_num > 1:
            selects = component_info['Selects']
            for i in range(select_num):
                IF_STATEMENT = "if( "
                FIRST = True
                for j in range(select_num):
                    if not FIRST:
                        IF_STATEMENT += " && "
                    else:
                        FIRST = not FIRST
                        
                    if component_info['IS_REVERSE']:
                        IF_STATEMENT += "(!"
                    else:
                        IF_STATEMENT += "("
                    
                    if i == j:
                        IF_STATEMENT += "tr.{} == 1)".format(selects[j])
                    else:
                        IF_STATEMENT += "tr.{} == 0)".format(selects[j])
                IF_STATEMENT += " ) begin"
                contents_append_tab_tab_endl(CONTENTS, IF_STATEMENT)
                
                input_signal = component_info['Inputs'][i]
                contents_append_tab_tab_endl(CONTENTS, TAB + '`uvm_info("selection", $sformatf("{} is {}, node[%0d] -> node[%0d] through {}", node_id, select_tree[{}].select[{}]), UVM_LOW)'.format(selects[j], i, input_signal, node_id, i))
                contents_append_tab_tab_endl(CONTENTS, TAB + "select_tree[{}].select_done[{}] = 1'b1;".format(node_id, i))
                contents_append_tab_tab_endl(CONTENTS, TAB + "return select_tree[{}].select[{}];".format(node_id, i))
                contents_append_tab_tab_endl(CONTENTS, "end")
                
        if select_num == 0:
            contents_append_tab_tab_endl(CONTENTS, '`uvm_info("selection", $sformatf("node[%0d] -> node[%0d]", node_id, select_tree[{}].select[0]), UVM_LOW)'.format(node_id))
            contents_append_tab_tab_endl(CONTENTS, "return select_tree[{}].select[0];".format(node_id))
        else:
            contents_append_tab_tab_endl(CONTENTS,  "return -1;")
        contents_append_tab_endl(CONTENTS, "end")
    contents_append_endl(CONTENTS, 'endfunction: selection')
    ###################################### tree selection #######################################
            
    contents_append_endl(CONTENTS, '')
    
    contents_append_endl(CONTENTS, '`endif // AUDIO_DATA_PATH_TRANSITION_MODEL__SV')
    
    if output_enable:
        write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')
    
    return components_info

def construct_components_info_sheet(components_info, wb, sheet_name = ''):
    if not sheet_name:
        ws = wb.active
    else:
        wb.create_sheet(sheet_name)
        ws = wb[sheet_name]

    col_names = ['Node_id', 'Type']
    rows = []
    max_inputs_num  = -1
    max_outputs_num = -1
    max_selects_num = -1
    max_select_num  = -1

    for component_info in components_info:
        max_inputs_num  = len(component_info['Inputs'])  if max_inputs_num < len(component_info['Inputs'])  else max_inputs_num
        max_outputs_num = len(component_info['Outputs']) if max_outputs_num < len(component_info['Outputs']) else max_outputs_num
        max_selects_num = len(component_info['Selects']) if max_selects_num < len(component_info['Selects']) else max_selects_num
        max_select_num  = len(component_info['select'])  if max_select_num < len(component_info['select'])  else max_select_num
        
    for i in range(max_inputs_num):
        col_names.append('Input[{}]'.format(i))
    
    for i in range(max_outputs_num):
        col_names.append('Output[{}]'.format(i))
    
    for i in range(max_selects_num):
        col_names.append('Select[{}]'.format(i))
        
    for i in range(max_select_num):
        col_names.append('Connect[{}]'.format(i))
    
    col_names.append('Default_value_is_high')
        
    for component_info in components_info:
        row = []
        row.append(component_info['NODE_ID'])
        row.append(component_info['Type'])
        for i in range(max_inputs_num):
            try:
                row.append(component_info['Inputs'][i])
            except:
                row.append(' ')
                
        for i in range(max_outputs_num):
            try:
                row.append(component_info['Outputs'][i])
            except:
                row.append(' ')
                
        for i in range(max_selects_num):
            try:
                row.append(component_info['Selects'][i])
            except:
                row.append(' ')
                
        for i in range(max_select_num):
            try:
                row.append(component_info['select'][i])
            except:
                row.append(' ')                
                
        row.append(component_info['IS_REVERSE'])
        
        rows.append(row)
        
    ws = create_excel_table_title(ws, col_names)

    for row in rows:
        ws.append(row)
        
    ws = excel_auto_width(ws)
    
    return ws

def output_components_info_xlsx(components_info, output_file_name = 'components_info.xlsx', output_dir = DATA_DIR):
    wb = Workbook()
    construct_components_info_sheet(components_info, wb)
    
    output_file_name = file_with_dir_path(output_file_name, output_dir)
    wb.save(output_file_name)

def get_select_signals_and_ports(components_info):
    ports   = []
    signals = []
    for component_info in components_info:
        input_num  = len(component_info['Inputs'])
        select_num = len(component_info['Selects'])
         
        if select_num > 0 and input_num > select_num:
            signal = component_info['Selects'][0]
            port_upper_bound = int(math.log(input_num, 2))-1 if check_pow_of_2(input_num) else int(math.log(input_num, 2))
            
            if port_upper_bound > 0:
                port   = 'logic [{}:0]{}'.format(port_upper_bound, signal)
            else:
                port   = 'logic {}'.format(signal)
                
            signals.append(signal)
            ports.append(port)
        else:
            for signal in component_info['Selects']:
                port = 'logic {}'.format(signal)
                signals.append(signal)
                ports.append(port)
                
    ports   = sorted(list(set(ports)))
    signals = sorted(list(set(signals)))
    
    return signals, ports

def gen_dut_wrapper(components_info, output_file_name = 'dut_wrapper.sv', output_dir = CONTENT_DIR):
    CONTENTS = []
    contents_append_endl(CONTENTS, SIGNATURE)
    signals, ports = get_select_signals_and_ports(components_info)
    
    contents_append_endl(CONTENTS, '// ****** audio data path VIP connection start ****** //')
    
    contents_append_endl(CONTENTS, '// select signals connection')
    for signal in signals:
        contents_append_endl(CONTENTS, '//assign audio_data_path_dump_mux_if_0.{}'.format(signal).ljust(70) + ' = ;')
        
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '')
    
    contents_append_endl(CONTENTS, '// output data signals connection')
    for idx, output_node in enumerate(get_nodes_by_type(components_info, 'Output_Node')):
        contents_append_endl(CONTENTS, '//output: {}'.format(output_node['Inputs'][0]))
        contents_append_endl(CONTENTS, '//assign audio_data_path_dump_data_if_0_{}.rreq'.format(idx).ljust(70) + ' = ;')
        contents_append_endl(CONTENTS, '//assign audio_data_path_dump_data_if_0_{}.data'.format(idx).ljust(70) + ' = ;')
        contents_append_endl(CONTENTS, '')
        
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_system_base_test(components_info, output_file_name = 'system_base_test.sv', output_dir = CONTENT_DIR):
    CONTENTS = []
    contents_append_endl(CONTENTS, SIGNATURE)
    
    contents_append_tab_endl(CONTENTS, '// ****** audio data path VIP source info start ****** //')
   
    for idx, input_node in enumerate(get_nodes_by_type(components_info, 'Input_Node')):
        contents_append_tab_endl(CONTENTS, '//input: {}'.format(input_node['Outputs'][0]))
        contents_append_tab_endl(CONTENTS, 'sys_cfg.audio_data_path_cfg[0].add_perf_struct(%3s, );' % idx)
        contents_append_endl(CONTENTS, '')
    
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_interface(components_info, output_file_name = "audio_data_path_dump_mux_interface.sv", output_dir = ENV_DIR):
    CONTENTS  = []
    contents_append_endl(CONTENTS, SIGNATURE)

    tmp = '`ifndef AUDIO_DATA_PATH_DUMP_MUX_INTERFACE__SV\n`define AUDIO_DATA_PATH_DUMP_MUX_INTERFACE__SV'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')

    tmp = 'interface audio_data_path_dump_mux_interface (input logic clk);'
    contents_append_endl(CONTENTS, tmp)

    tmp = TAB + 'logic rst;'
    contents_append_endl(CONTENTS, tmp)

    signals, ports = get_select_signals_and_ports(components_info)
                
    for port in ports:
        contents_append_endl(CONTENTS, TAB + port + ';')
        
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '')

    tmp = TAB + "modport monitor("
    contents_append_endl(CONTENTS, tmp)

    for signal in signals:
        tmp = TAB + TAB + TAB + TAB + TAB + 'input {},'.format(signal)
        contents_append_endl(CONTENTS, tmp)
        
    CONTENTS[-1] = CONTENTS[-1].replace(',','')

    tmp = TAB + "              );"
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')

    tmp = 'endinterface: audio_data_path_dump_mux_interface'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')

    tmp = '`endif // AUDIO_DATA_PATH_DUMP_MUX_INTERFACE__SV'
    contents_append_endl(CONTENTS, tmp)

    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_transaction(components_info, output_file_name = "cust_audio_data_path_dump_mux_transaction.sv", output_dir = ENV_DIR):
    CONTENTS  = []
    contents_append_endl(CONTENTS, SIGNATURE)

    tmp = '`ifndef CUST_AUDIO_DATA_PATH_DUMP_MUX_TRANSACTION__SV\n`define CUST_AUDIO_DATA_PATH_DUMP_MUX_TRANSACTION__SV'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')

    tmp = 'class cust_audio_data_path_dump_mux_transaction extends audio_data_path_dump_mux_transaction;'
    contents_append_endl(CONTENTS, tmp)

    signals, ports = get_select_signals_and_ports(components_info)
    
    for port in ports:
        tmp = TAB + port + ';'
        contents_append_endl(CONTENTS, tmp)
        
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '')

    tmp = TAB + "`uvm_object_utils_begin(cust_audio_data_path_dump_mux_transaction)"
    contents_append_endl(CONTENTS, tmp)

    for signal in signals:
        tmp = TAB + TAB + '`uvm_field_int({}, UVM_DEFAULT)'.format(signal)
        contents_append_endl(CONTENTS, tmp)
            
    tmp = TAB + "`uvm_object_utils_end"
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')

    tmp = TAB + 'extern function new(string name = "cust_audio_data_path_dump_mux_transaction");'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = 'endclass: cust_audio_data_path_dump_mux_transaction'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = 'function cust_audio_data_path_dump_mux_transaction::new(string name = "cust_audio_data_path_dump_mux_transaction");'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = TAB + "super.new(name);"
    contents_append_endl(CONTENTS, tmp)
    
    tmp = "endfunction: new"
    contents_append_endl(CONTENTS, tmp)

    tmp = '`endif // CUST_AUDIO_DATA_PATH_DUMP_MUX_TRANSACTION__SV'
    contents_append_endl(CONTENTS, tmp)

    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_monitor(components_info, output_file_name = "cust_audio_data_path_dump_mux_monitor.sv", output_dir = ENV_DIR):
    CONTENTS  = []
    contents_append_endl(CONTENTS, SIGNATURE)

    tmp = '`ifndef CUST_AUDIO_DATA_PATH_DUMP_MUX_MONITOR__SV\n`define CUST_AUDIO_DATA_PATH_DUMP_MUX_MONITOR__SV'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')

    tmp = 'class cust_audio_data_path_dump_mux_monitor extends audio_data_path_dump_mux_monitor #(cust_audio_data_path_dump_mux_transaction);'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = TAB + 'virtual audio_data_path_dump_mux_interface audio_data_path_dump_mux_if;'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = TAB + '`uvm_component_utils(cust_audio_data_path_dump_mux_monitor)'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = TAB + 'extern function new(string name = "cust_audio_data_path_dump_mux_monitor", uvm_component parent);'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = TAB + 'extern virtual function void build_phase(uvm_phase phase);'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = TAB + 'extern protected virtual task sample_mux(ref cust_audio_data_path_dump_mux_transaction tr);'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = 'endclass: cust_audio_data_path_dump_mux_monitor'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = 'function cust_audio_data_path_dump_mux_monitor::new(string name = "cust_audio_data_path_dump_mux_monitor", uvm_component parent);'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = TAB + 'super.new(name, parent);'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = 'endfunction: new'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = 'function void cust_audio_data_path_dump_mux_monitor::build_phase(uvm_phase phase);'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = TAB + 'super.build_phase(phase);'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = TAB + 'if (!uvm_config_db#(virtual audio_data_path_dump_mux_interface)::get(this, "", "audio_data_path_dump_mux_if", audio_data_path_dump_mux_if))'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = TAB + TAB + '`uvm_fatal("NOVIF", {"audio_data_path_dump_mux_if must be set for ", get_full_name()})'
    contents_append_endl(CONTENTS, tmp)
    
    tmp = 'endfunction: build_phase'
    contents_append_endl(CONTENTS, tmp)
    contents_append_endl(CONTENTS, '')
    
    tmp = 'task cust_audio_data_path_dump_mux_monitor::sample_mux(ref cust_audio_data_path_dump_mux_transaction tr);'
    contents_append_endl(CONTENTS, tmp)
        
    signals, ports = get_select_signals_and_ports(components_info)
        
    for signal in signals:
        tmp = TAB + 'tr.{} = audio_data_path_dump_mux_if.{};'.format(signal, signal)
        contents_append_endl(CONTENTS, tmp)
    
    tmp = 'endtask: sample_mux'
    contents_append_endl(CONTENTS, tmp)

    tmp = '`endif // CUST_AUDIO_DATA_PATH_DUMP_MUX_MONITOR__SV'
    contents_append_endl(CONTENTS, tmp)

    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

def gen_coverage(components_info, output_file_name = 'cust_audio_data_path_coverage_model.sv', output_dir = ENV_DIR):
    CONTENTS  = []
    contents_append_endl(CONTENTS, SIGNATURE) 
    contents_append_endl(CONTENTS, '`ifndef CUST_AUDIO_DATA_PATH_COVERAGE_MODEL__SV')
    contents_append_endl(CONTENTS, '`define CUST_AUDIO_DATA_PATH_COVERAGE_MODEL__SV')
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, 'class cust_audio_data_path_coverage_model extends audio_data_path_coverage_model #(cust_audio_data_path_dump_mux_transaction, cust_audio_data_path_transition_model);')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, '`uvm_component_utils(cust_audio_data_path_coverage_model)')
    contents_append_endl(CONTENTS, '')
    contents_append_tab_endl(CONTENTS, 'covergroup cg_components;')
    contents_append_tab_tab_endl(CONTENTS, 'option.per_instance = 1;')
    
    for component_info in components_info:
        if component_info['Type'] != 'Input_Node' and len(component_info['Selects']) > 0:
            NODE_ID = component_info['NODE_ID']
            coverpoint_name = '{}_COMPONENT_{}'.format(component_info['Type'].replace('/','_'), NODE_ID)
            contents_append_tab_tab_endl(CONTENTS, coverpoint_name + ':')
            COVERPOINT  = TAB + 'coverpoint '
            COVERPOINT += ' {\n'
            for idx, input in enumerate(component_info['Inputs']):
                COVERPOINT += TAB + TAB + TAB + TAB + TAB + TAB + TAB + 'transition.select_tree[{}].select_done[{}]'.format(NODE_ID, idx)
                if idx == len(component_info['Inputs']) -1:
                    COVERPOINT += '\n' + TAB + TAB + TAB + TAB + TAB + TAB + '}'
                else:
                    COVERPOINT += ',\n'
            
            COVERPOINT += '\n' + TAB + TAB + TAB + TAB + TAB + TAB + '{\n'
            for idx, input in enumerate(component_info['Inputs']):
                COVERPOINT += TAB + TAB + TAB + TAB + TAB + TAB + TAB + 'bins {}_Output_select_{} = '.format(coverpoint_name, idx)
                COVERPOINT += '{'
                COVERPOINT += "{}'b".format(len(component_info['Inputs']))
                for i in range(len(component_info['Inputs'])):
                    if i == idx:
                        COVERPOINT += '1'
                    else:
                        COVERPOINT += '0'
                COVERPOINT += '};\n'
            COVERPOINT += TAB + TAB + TAB + TAB + TAB + TAB + '}'
            contents_append_tab_tab_endl(CONTENTS, COVERPOINT)
    contents_append_tab_endl(CONTENTS, 'endgroup')
    contents_append_endl(CONTENTS, '')
        
    contents_append_tab_endl(CONTENTS, 'extern function new(string name = "cust_audio_data_path_coverage_model", uvm_component parent);')
    contents_append_tab_endl(CONTENTS, 'extern virtual function void sample_cov();')
    contents_append_tab_endl(CONTENTS, 'extern virtual function real get_cov();')
    contents_append_endl(CONTENTS, 'endclass: cust_audio_data_path_coverage_model')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '/** new function for cust_audio_data_path_coverage_model */')
    contents_append_endl(CONTENTS, 'function cust_audio_data_path_coverage_model::new(string name = "cust_audio_data_path_coverage_model", uvm_component parent);')
    contents_append_tab_endl(CONTENTS, 'super.new(name, parent);')
    contents_append_tab_endl(CONTENTS, 'cg_components = new();')
    contents_append_endl(CONTENTS, 'endfunction: new')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '/** sample_cov for cust_audio_data_path_coverage_model */')
    contents_append_endl(CONTENTS, 'function void cust_audio_data_path_coverage_model::sample_cov();')
    contents_append_tab_endl(CONTENTS, 'cg_components.sample();')
    contents_append_endl(CONTENTS, 'endfunction: sample_cov')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '/** get_cov for cust_audio_data_path_coverage_model */')
    contents_append_endl(CONTENTS, 'function real cust_audio_data_path_coverage_model::get_cov();')
    contents_append_tab_endl(CONTENTS, 'get_cov = cg_components.get_inst_coverage();')
    contents_append_endl(CONTENTS, 'endfunction: get_cov')
    
    contents_append_endl(CONTENTS, '')
    contents_append_endl(CONTENTS, '`endif // CUST_AUDIO_DATA_PATH_COVERAGE_MODEL__SV')
    
    write_contents_to_file(CONTENTS, output_file_name, output_dir, ENDL = '')

if __name__ == '__main__':
    ################################################################## main ##################################################################
    components_parsing_rule = get_components_parsing_rule()
    components_info = get_components_info(components_parsing_rule)
    components_info = make_distinct(components_info)
    components_info = to_lower(components_info)
    components_info = sort_components_info(components_info)
    components_info = set_components_id(components_info)

    print('output_components_info...')
    output_components_info(components_info)
    print('done\n')

    print('parse_register_file...')
    all_register_info = parse_register_file()
    if not all_register_info:
        ERROR_MESSAGE('all_register_info not prepared')
    else:
        output_register_info(all_register_info)
        print('done\n')

    ####################################### check flow #######################################
    print('check_single_output...')
    CHECK_PASS = check_single_output(components_info)
    CHECK_ALL_PASS = check_merge(CHECK_PASS, CHECK_PASS)

    print('check_output_distinct...')
    CHECK_PASS = check_output_distinct(components_info)
    CHECK_ALL_PASS = check_merge(CHECK_ALL_PASS, CHECK_PASS)

    print('check_output_connection...')
    CHECK_PASS = check_output_connection(components_info)
    CHECK_ALL_PASS = check_merge(CHECK_ALL_PASS, CHECK_PASS)

    print('check_input_connection...')
    CHECK_PASS = check_input_connection(components_info)
    CHECK_ALL_PASS = check_merge(CHECK_ALL_PASS, CHECK_PASS)

    print('check_select_and_input_num...')
    CHECK_PASS = check_select_and_input_num(components_info)
    CHECK_ALL_PASS = check_merge(CHECK_ALL_PASS, CHECK_PASS)

    print('check_select_in_all_register_info...')
    CHECK_PASS = check_select_in_all_register_info(components_info, all_register_info)
    CHECK_ALL_PASS = check_merge(CHECK_ALL_PASS, CHECK_PASS)

    print('check_select_distinct...')
    CHECK_PASS = check_select_distinct(components_info)
    CHECK_PASS = True
    CHECK_ALL_PASS = check_merge(CHECK_ALL_PASS, CHECK_PASS)

    print('check_components_info_and_all_register_info...')
    CHECK_PASS = check_components_info_and_all_register_info(components_info, all_register_info)
    CHECK_ALL_PASS = check_merge(CHECK_ALL_PASS, CHECK_PASS)
    ####################################### check flow #######################################
    
    print('output_mux_gentop...')
    output_mux_gentop(components_info)
    print('done\n')

    if CHECK_ALL_PASS:
        PASS_MESSAGE('Check all pass\n')
        
        if not FOR_SD_CHECK_ONLY:  
            print('make_legal...')
            components_info = make_legal(components_info)
            print('done\n')
        else:
            print('make_legal...')
            components_info = make_legal(components_info)
            print('done\n')
                
        if not FOR_SD_CHECK_ONLY:
            print('gen_content_of_cust_system_configuration...')
            gen_content_of_cust_system_configuration(components_info)
            print('done\n')
            
            print('gen_audio_data_path_golden_pattern...')
            gen_audio_data_path_golden_pattern(components_info)
            gen_audio_data_phase_golden_pattern(components_info)
            print('done\n')
            
        if not FOR_SD_CHECK_ONLY:
            print('gen_transition_model...')
            components_info = gen_transition_model(components_info)
            print('done\n')
        else:
            components_info = gen_transition_model(components_info, output_enable = False)
            
        print('output_components_info...')
        output_components_info(components_info)
        print('done\n')
        

            
        print('output_components_info_xlsx...')
        output_components_info_xlsx(components_info)
        print('done\n')
            
        if not FOR_SD_CHECK_ONLY:   
            print('gen_dut_wrapper...')
            gen_dut_wrapper(components_info)
            print('done\n')
            
            print('gen_system_base_test...')
            gen_system_base_test(components_info)
            print('done\n')
            
            print('gen_interface...')
            gen_interface(components_info)
            print('done\n')
            
            print('gen_transaction...')
            gen_transaction(components_info)
            print('done\n')
            
            print('gen_monitor...')
            gen_monitor(components_info)
            print('done\n')
            
            print('gen_coverage...')
            gen_coverage(components_info)
            print('done\n')

    else:
        FAIL_MESSAGE('check failed, please check visio file')